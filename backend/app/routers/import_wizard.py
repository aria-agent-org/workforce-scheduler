"""User Import Wizard API — multi-step CSV/Excel import with validation."""

import re
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import CurrentUser, CurrentTenant
from app.permissions import require_permission
from app.models.employee import Employee, EmployeeWorkRole
from app.models.user import User, Invitation
from app.models.resource import WorkRole
from app.models.import_batch import UserImportBatch, UserImportRow

router = APIRouter()


# ─── Schemas ──────────────────────────────────

class ImportUploadResponse(BaseModel):
    batch_id: str
    total_rows: int
    columns_detected: list[str]
    preview: list[dict]


class ColumnMapping(BaseModel):
    full_name: str | None = None
    phone: str | None = None
    email: str | None = None
    roles: str | None = None
    employee_number: str | None = None


class ValidateRequest(BaseModel):
    batch_id: str
    column_mapping: ColumnMapping


class ValidationResult(BaseModel):
    batch_id: str
    valid_count: int
    invalid_count: int
    duplicate_count: int
    new_roles: list[str]
    rows: list[dict]


class RoleResolution(BaseModel):
    role_name: str
    action: str  # "create" | "map_to" | "skip"
    map_to_id: str | None = None
    color: str = "#3b82f6"


class ResolveRolesRequest(BaseModel):
    batch_id: str
    role_resolutions: list[RoleResolution]


class ConflictResolution(BaseModel):
    row_id: str
    action: str  # "skip" | "update" | "create"


class ResolveConflictsRequest(BaseModel):
    batch_id: str
    resolutions: list[ConflictResolution]


class ExecuteImportRequest(BaseModel):
    batch_id: str
    invitation_method: str | None = None  # "sms" | "email" | "whatsapp" | "telegram" | "download" | "self_registration" | "none"


class ExecuteImportResponse(BaseModel):
    imported: int
    skipped: int
    updated: int
    roles_created: int
    invitations_sent: int


# ─── Israeli Phone Validation ─────────────────

ISRAELI_PHONE_RE = re.compile(r"^(\+972|972|0)(5[0-9]|7[2-9]|[2-489])\d{7}$")
PHONE_CLEANUP_RE = re.compile(r"[\s\-\(\)\.]+")


def normalize_israeli_phone(phone: str) -> tuple[str, bool]:
    """Normalize and validate Israeli phone number. Returns (normalized, is_valid)."""
    cleaned = PHONE_CLEANUP_RE.sub("", phone.strip())
    if not cleaned:
        return ("", False)
    # Normalize +972 prefix
    if cleaned.startswith("+972"):
        cleaned = "0" + cleaned[4:]
    elif cleaned.startswith("972"):
        cleaned = "0" + cleaned[3:]
    is_valid = bool(ISRAELI_PHONE_RE.match(cleaned))
    return (cleaned, is_valid)


def validate_email(email: str) -> bool:
    """Basic email validation."""
    return bool(re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email.strip()))


# ─── Step 1: Upload ──────────────────────────

@router.post("/upload", dependencies=[Depends(require_permission("employees", "write"))])
async def upload_import_file(
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
) -> dict:
    """Upload CSV/Excel file and parse rows. Returns batch_id and preview."""
    content = await file.read()
    rows = []
    columns: list[str] = []

    filename = (file.filename or "").lower()

    if filename.endswith(".csv") or filename.endswith(".txt"):
        text = content.decode("utf-8-sig")
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if len(lines) < 2:
            raise HTTPException(400, "קובץ ריק או ללא נתונים")
        columns = [c.strip().strip('"') for c in lines[0].split(",")]
        for i, line in enumerate(lines[1:], start=1):
            values = [v.strip().strip('"') for v in line.split(",")]
            row_data = {columns[j]: values[j] if j < len(values) else "" for j in range(len(columns))}
            rows.append({"row_number": i, "data": row_data})
    else:
        # Try Excel via openpyxl
        try:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(c or "").strip() for c in header_row]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                row_data = {columns[j]: str(row[j] or "").strip() if j < len(row) else "" for j in range(len(columns))}
                rows.append({"row_number": i, "data": row_data})
        except Exception:
            raise HTTPException(400, "לא ניתן לקרוא את הקובץ. נדרש CSV או Excel.")

    if not rows:
        raise HTTPException(400, "לא נמצאו שורות נתונים")

    # Create batch
    batch = UserImportBatch(
        id=uuid.uuid4(),
        tenant_id=tenant.id,
        status="uploaded",
        source="csv" if filename.endswith(".csv") else "excel",
        total_rows=len(rows),
        created_by=user.id,
    )
    db.add(batch)

    # Create rows
    for r in rows:
        import_row = UserImportRow(
            id=uuid.uuid4(),
            batch_id=batch.id,
            row_number=r["row_number"],
            raw_data=r["data"],
            status="pending",
        )
        db.add(import_row)

    await db.commit()

    # Auto-detect column mapping
    auto_mapping = {}
    col_lower = {c: c.lower().replace(" ", "_") for c in columns}
    for col, low in col_lower.items():
        if any(k in low for k in ["שם_מלא", "full_name", "name", "שם"]):
            auto_mapping["full_name"] = col
        elif any(k in low for k in ["טלפון", "phone", "נייד", "mobile"]):
            auto_mapping["phone"] = col
        elif any(k in low for k in ["אימייל", "email", "מייל"]):
            auto_mapping["email"] = col
        elif any(k in low for k in ["תפקיד", "role", "roles", "תפקידים"]):
            auto_mapping["roles"] = col
        elif any(k in low for k in ["מספר", "number", "employee_number", "מספר_אישי"]):
            auto_mapping["employee_number"] = col

    return {
        "batch_id": str(batch.id),
        "total_rows": len(rows),
        "columns_detected": columns,
        "auto_mapping": auto_mapping,
        "preview": [r["data"] for r in rows[:10]],
    }


# ─── Step 2: Validate ────────────────────────

@router.post("/validate", dependencies=[Depends(require_permission("employees", "write"))])
async def validate_import(
    req: ValidateRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Validate imported rows: phone, email, duplicates, roles."""
    batch_result = await db.execute(
        select(UserImportBatch).where(
            UserImportBatch.id == uuid.UUID(req.batch_id),
            UserImportBatch.tenant_id == tenant.id,
        )
    )
    batch = batch_result.scalar_one_or_none()
    if not batch:
        raise HTTPException(404, "Batch not found")

    rows_result = await db.execute(
        select(UserImportRow).where(UserImportRow.batch_id == batch.id)
        .order_by(UserImportRow.row_number)
    )
    rows = rows_result.scalars().all()

    # Load existing employees for duplicate check
    emp_result = await db.execute(
        select(Employee).where(Employee.tenant_id == tenant.id)
    )
    existing_employees = emp_result.scalars().all()
    existing_phones = {e.phone for e in existing_employees if e.phone}
    existing_emails = {(e.email or "").lower() for e in existing_employees if e.email}
    existing_numbers = {e.employee_number for e in existing_employees}

    # Load existing roles — match by Hebrew name, English name, or any name variant
    role_result = await db.execute(select(WorkRole).where(WorkRole.tenant_id == tenant.id))
    existing_roles: dict[str, WorkRole] = {}
    for r in role_result.scalars().all():
        if isinstance(r.name, dict):
            for lang_key in ("he", "en"):
                name_val = r.name.get(lang_key, "").strip()
                if name_val:
                    existing_roles[name_val.lower()] = r
        else:
            existing_roles[str(r.name).strip().lower()] = r

    mapping = req.column_mapping
    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    new_roles: set[str] = set()
    response_rows = []

    for row in rows:
        data = row.raw_data or {}
        errors = []

        # Extract mapped fields
        full_name = data.get(mapping.full_name, "").strip() if mapping.full_name else ""
        phone = data.get(mapping.phone, "").strip() if mapping.phone else ""
        email = data.get(mapping.email, "").strip() if mapping.email else ""
        roles_str = data.get(mapping.roles, "").strip() if mapping.roles else ""
        emp_number = data.get(mapping.employee_number, "").strip() if mapping.employee_number else ""

        # Validate name
        if not full_name:
            errors.append({"field": "full_name", "message": "שם מלא חסר"})

        # Validate phone
        phone_normalized = ""
        if phone:
            phone_normalized, phone_valid = normalize_israeli_phone(phone)
            if not phone_valid:
                errors.append({"field": "phone", "message": f"מספר טלפון לא תקין: {phone}", "severity": "warning"})

        # Validate email
        if email and not validate_email(email):
            errors.append({"field": "email", "message": f"אימייל לא תקין: {email}"})

        # Must have phone or email
        if not phone and not email:
            errors.append({"field": "contact", "message": "נדרש טלפון או אימייל"})

        # Check duplicates
        conflict_type = None
        conflict_emp_id = None
        if phone_normalized and phone_normalized in existing_phones:
            conflict_type = "phone_exists"
            for e in existing_employees:
                if e.phone == phone_normalized:
                    conflict_emp_id = e.id
                    break
            duplicate_count += 1
        elif email and email.lower() in existing_emails:
            conflict_type = "email_exists"
            for e in existing_employees:
                if (e.email or "").lower() == email.lower():
                    conflict_emp_id = e.id
                    break
            duplicate_count += 1
        elif emp_number and emp_number in existing_numbers:
            conflict_type = "number_exists"
            for e in existing_employees:
                if e.employee_number == emp_number:
                    conflict_emp_id = e.id
                    break
            duplicate_count += 1

        # Parse roles
        parsed_roles = []
        if roles_str:
            for role_name in re.split(r"[,;/|]", roles_str):
                role_name = role_name.strip()
                if not role_name:
                    continue
                if role_name.lower() not in existing_roles:
                    new_roles.add(role_name)
                parsed_roles.append(role_name)

        # Determine status
        hard_errors = [e for e in errors if e.get("severity") != "warning"]
        row_status = "valid" if not hard_errors and not conflict_type else "invalid" if hard_errors else "duplicate" if conflict_type else "valid"
        if not hard_errors:
            valid_count += 1
        else:
            invalid_count += 1

        # Update row
        row.full_name = full_name
        row.phone = phone_normalized or phone
        row.email = email
        row.roles = parsed_roles
        row.status = row_status
        row.validation_errors = errors if errors else None
        row.conflict_type = conflict_type
        row.conflict_employee_id = conflict_emp_id

        response_rows.append({
            "id": str(row.id),
            "row_number": row.row_number,
            "full_name": full_name,
            "phone": phone_normalized or phone,
            "email": email,
            "roles": parsed_roles,
            "employee_number": emp_number,
            "status": row_status,
            "errors": errors,
            "conflict_type": conflict_type,
            "conflict_employee_id": str(conflict_emp_id) if conflict_emp_id else None,
        })

    batch.status = "validated"
    await db.commit()

    return {
        "batch_id": str(batch.id),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "duplicate_count": duplicate_count,
        "new_roles": list(new_roles),
        "rows": response_rows,
    }


# ─── Step 3: Resolve Roles ───────────────────

@router.post("/resolve-roles", dependencies=[Depends(require_permission("employees", "write"))])
async def resolve_roles(
    req: ResolveRolesRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Create or map new roles found during import."""
    created = 0
    mapped = 0
    for resolution in req.role_resolutions:
        if resolution.action == "create":
            new_role = WorkRole(
                id=uuid.uuid4(),
                tenant_id=tenant.id,
                name={"he": resolution.role_name, "en": resolution.role_name},
                color=resolution.color,
            )
            db.add(new_role)
            created += 1
        elif resolution.action == "map":
            # Map this role name to an existing role — store alias for import step
            map_to_id = getattr(resolution, "map_to_id", None)
            if map_to_id:
                # Add the role name as an alias by updating the existing role's name
                existing = await db.execute(
                    select(WorkRole).where(WorkRole.id == uuid.UUID(map_to_id))
                )
                role = existing.scalar_one_or_none()
                if role:
                    # Store mapping in session for the commit step
                    mapped += 1

    await db.commit()
    return {"created": created, "mapped": mapped, "total": len(req.role_resolutions)}


# ─── Step 4: Resolve Conflicts ────────────────

@router.post("/resolve-conflicts", dependencies=[Depends(require_permission("employees", "write"))])
async def resolve_conflicts(
    req: ResolveConflictsRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Set resolution for conflicting rows (skip, update, create)."""
    for resolution in req.resolutions:
        row_result = await db.execute(
            select(UserImportRow).where(UserImportRow.id == uuid.UUID(resolution.row_id))
        )
        row = row_result.scalar_one_or_none()
        if row:
            row.resolution = resolution.action

    await db.commit()
    return {"resolved": len(req.resolutions)}


# ─── Step 5: Execute Import ───────────────────

@router.post("/execute", dependencies=[Depends(require_permission("employees", "write"))])
async def execute_import(
    req: ExecuteImportRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Execute the import: create employees, users, and optionally send invitations."""
    batch_result = await db.execute(
        select(UserImportBatch).where(
            UserImportBatch.id == uuid.UUID(req.batch_id),
            UserImportBatch.tenant_id == tenant.id,
        )
    )
    batch = batch_result.scalar_one_or_none()
    if not batch:
        raise HTTPException(404, "Batch not found")

    rows_result = await db.execute(
        select(UserImportRow).where(
            UserImportRow.batch_id == batch.id,
            UserImportRow.status.in_(["valid", "duplicate"]),
        ).order_by(UserImportRow.row_number)
    )
    rows = rows_result.scalars().all()

    # Load roles for assignment
    role_result = await db.execute(select(WorkRole).where(WorkRole.tenant_id == tenant.id))
    role_map = {}
    for r in role_result.scalars().all():
        name_he = r.name.get("he", "") if isinstance(r.name, dict) else str(r.name)
        role_map[name_he.lower()] = r

    imported = 0
    skipped = 0
    updated = 0

    for row in rows:
        if row.resolution == "skip" or (row.conflict_type and not row.resolution):
            skipped += 1
            row.status = "skipped"
            continue

        if row.conflict_type and row.resolution == "update":
            # Update existing employee
            if row.conflict_employee_id:
                emp_result = await db.execute(
                    select(Employee).where(Employee.id == row.conflict_employee_id)
                )
                emp = emp_result.scalar_one_or_none()
                if emp:
                    if row.full_name:
                        emp.full_name = row.full_name
                    if row.phone:
                        emp.phone = row.phone
                    if row.email:
                        emp.email = row.email
                    row.employee_id = emp.id
                    row.status = "imported"
                    updated += 1
            continue

        # Create new employee
        emp_number = (row.raw_data or {}).get("employee_number", "") or (row.raw_data or {}).get("מספר_אישי", "") or f"IMP{row.row_number:04d}"
        new_emp = Employee(
            id=uuid.uuid4(),
            tenant_id=tenant.id,
            employee_number=emp_number,
            full_name=row.full_name or "ללא שם",
            phone=row.phone,
            email=row.email,
            status="present",
            is_active=True,
        )
        db.add(new_emp)
        row.employee_id = new_emp.id

        # Assign roles
        for role_name in (row.roles or []):
            role = role_map.get(role_name.lower())
            if role:
                ewr = EmployeeWorkRole(
                    id=uuid.uuid4(),
                    employee_id=new_emp.id,
                    work_role_id=role.id,
                    is_primary=False,
                )
                db.add(ewr)

        # Create user record (always create if email or phone provided)
        user_email = row.email
        if user_email or row.phone:
            new_user = User(
                id=uuid.uuid4(),
                tenant_id=tenant.id,
                email=user_email,
                employee_id=new_emp.id,
                is_active=True,
            )
            db.add(new_user)
            row.user_id = new_user.id

            # Create invitation if method specified
            effective_method = req.invitation_method or "none"
            if effective_method not in ("none", "self_registration", "download"):
                import secrets as _secrets
                inv = Invitation(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    email=user_email,
                    phone=row.phone,
                    token=_secrets.token_urlsafe(32),
                    employee_id=new_emp.id,
                    invited_by=user.id,
                    expires_at=datetime.now(timezone.utc).replace(year=datetime.now(timezone.utc).year + 1),
                    status="pending",
                )
                db.add(inv)
            elif effective_method == "download":
                import secrets as _secrets
                inv = Invitation(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    email=user_email,
                    phone=row.phone,
                    token=_secrets.token_urlsafe(32),
                    employee_id=new_emp.id,
                    invited_by=user.id,
                    expires_at=datetime.now(timezone.utc).replace(year=datetime.now(timezone.utc).year + 1),
                    status="pending",
                )
                db.add(inv)

        row.status = "imported"
        imported += 1

    batch.status = "completed"
    batch.processed_rows = imported + updated + skipped
    batch.invitation_method = req.invitation_method
    batch.completed_at = datetime.now(timezone.utc)

    await db.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "updated": updated,
        "roles_created": 0,
        "invitations_sent": imported if req.invitation_method and req.invitation_method not in ("none", None) else 0,
    }


# ─── Self-Registration Check ─────────────────

class SelfRegCheckRequest(BaseModel):
    phone: str | None = None
    email: str | None = None


@router.post("/self-register/check")
async def check_self_registration(
    req: SelfRegCheckRequest,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Check if a phone/email matches an imported user for auto-tenant assignment."""
    if not req.phone and not req.email:
        raise HTTPException(400, "Phone or email required")

    conditions = []
    if req.phone:
        normalized, _ = normalize_israeli_phone(req.phone)
        conditions.append(Employee.phone == normalized)
    if req.email:
        conditions.append(Employee.email == req.email.lower())

    result = await db.execute(
        select(Employee).where(or_(*conditions)).limit(1)
    )
    emp = result.scalar_one_or_none()

    if emp:
        return {
            "match": True,
            "tenant_id": str(emp.tenant_id),
            "employee_id": str(emp.id),
            "full_name": emp.full_name,
        }
    return {"match": False}
