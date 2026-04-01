"""Scheduling endpoints: windows, missions, templates, assignments, swaps."""

from datetime import date, datetime, time, timedelta
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import Response as FastAPIResponse
from pydantic import BaseModel as PydanticBaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import CurrentUser, CurrentTenant
from app.permissions import require_permission
from app.websockets.manager import manager as ws_manager
from app.models.scheduling import (
    ScheduleWindow, ScheduleWindowEmployee, ScheduleWindowLifecycleEvent,
    MissionType, MissionTemplate,
    Mission, MissionAssignment, SwapRequest, DailyBoardTemplate,
)
from app.models.employee import Employee, EmployeeWorkRole
from app.models.audit import AuditLog
from app.schemas.scheduling import (
    ScheduleWindowCreate, ScheduleWindowUpdate, ScheduleWindowEmployeeAdd,
    MissionTypeCreate, MissionTypeUpdate, MissionTypeResponse,
    MissionTemplateCreate, MissionTemplateUpdate, MissionTemplateResponse,
    MissionCreate, MissionUpdate, MissionGenerateRequest,
    MissionAssignmentCreate, SwapRequestCreate,
)
from app.schemas.jsonb_validators import MissionSlot, RecurrencePattern, TimelineItem
from pydantic import ValidationError as PydanticValidationError

router = APIRouter()


def _validate_jsonb(items: list | None, model_cls: type, field_name: str) -> None:
    """Validate a list of JSONB items against a Pydantic model. Raises HTTPException 422."""
    if items is None:
        return
    if not isinstance(items, list):
        raise HTTPException(status_code=422, detail=f"שדה {field_name} חייב להיות רשימה")
    errors = []
    for i, item in enumerate(items):
        try:
            model_cls.model_validate(item)
        except PydanticValidationError as exc:
            for err in exc.errors():
                errors.append(f"{field_name}[{i}].{'.'.join(str(l) for l in err['loc'])}: {err['msg']}")
    if errors:
        raise HTTPException(status_code=422, detail={"message": "שגיאת אימות נתונים", "errors": errors})


def _validate_recurrence(rec: dict | None) -> None:
    """Validate recurrence JSONB against RecurrencePattern."""
    if rec is None:
        return
    try:
        RecurrencePattern.model_validate(rec)
    except PydanticValidationError as exc:
        errors = [f"{'.'.join(str(l) for l in e['loc'])}: {e['msg']}" for e in exc.errors()]
        raise HTTPException(status_code=422, detail={"message": "שגיאת אימות תבנית חזרה", "errors": errors})


# ═══════════════════════════════════════════
# Schedule Windows
# ═══════════════════════════════════════════

@router.get("/schedule-windows")
async def list_schedule_windows(
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    status_filter: str | None = None,
) -> list[dict]:
    query = select(ScheduleWindow).where(ScheduleWindow.tenant_id == tenant.id)
    if status_filter:
        query = query.where(ScheduleWindow.status == status_filter)
    query = query.order_by(ScheduleWindow.start_date.desc())
    result = await db.execute(query)
    windows = result.scalars().all()
    items = []
    for w in windows:
        emp_count = (await db.execute(
            select(func.count()).where(ScheduleWindowEmployee.schedule_window_id == w.id)
        )).scalar() or 0
        items.append({
            "id": str(w.id),
            "tenant_id": str(w.tenant_id),
            "name": w.name,
            "start_date": str(w.start_date),
            "end_date": str(w.end_date),
            "status": w.status,
            "paused_at": str(w.paused_at) if w.paused_at else None,
            "notes": w.notes,
            "settings_override": w.settings_override,
            "employee_count": emp_count,
            "created_at": str(w.created_at),
            "updated_at": str(w.updated_at),
        })
    return items


@router.post("/schedule-windows", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permission("missions", "write"))])
async def create_schedule_window(
    data: ScheduleWindowCreate,
    tenant: CurrentTenant,
    user: CurrentUser,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    if data.end_date <= data.start_date:
        raise HTTPException(status_code=400, detail="תאריך סיום חייב להיות אחרי תאריך התחלה")
    window = ScheduleWindow(tenant_id=tenant.id, **data.model_dump())
    db.add(window)
    await db.flush()
    await db.refresh(window)
    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="create",
        entity_type="schedule_window", entity_id=window.id,
        after_state={"name": window.name, "status": window.status},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    return {
        "id": str(window.id), "tenant_id": str(window.tenant_id),
        "name": window.name, "start_date": str(window.start_date),
        "end_date": str(window.end_date), "status": window.status,
        "notes": window.notes, "employee_count": 0,
        "created_at": str(window.created_at), "updated_at": str(window.updated_at),
    }


@router.get("/schedule-windows/{window_id}")
async def get_schedule_window(
    window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    emp_count = (await db.execute(
        select(func.count()).where(ScheduleWindowEmployee.schedule_window_id == w.id)
    )).scalar() or 0
    return {
        "id": str(w.id), "tenant_id": str(w.tenant_id), "name": w.name,
        "start_date": str(w.start_date), "end_date": str(w.end_date),
        "status": w.status, "notes": w.notes, "employee_count": emp_count,
        "created_at": str(w.created_at), "updated_at": str(w.updated_at),
    }


@router.patch("/schedule-windows/{window_id}")
async def update_schedule_window(
    window_id: UUID, data: ScheduleWindowUpdate, tenant: CurrentTenant,
    user: CurrentUser, request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")

    # Concurrent editing protection — require version match
    update_data = data.model_dump(exclude_unset=True)
    client_version = update_data.pop("version", None)
    if client_version is not None and client_version != w.version:
        raise HTTPException(
            status_code=409,
            detail="מישהו אחר ערך את לוח העבודה. רענן ונסה שוב",
        )

    before = {"name": w.name, "status": w.status}
    for key, value in update_data.items():
        setattr(w, key, value)
    w.version = (w.version or 1) + 1
    await db.flush()
    await db.refresh(w)
    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="update",
        entity_type="schedule_window", entity_id=w.id,
        before_state=before, after_state={"name": w.name, "status": w.status},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    return {"id": str(w.id), "name": w.name, "status": w.status,
            "start_date": str(w.start_date), "end_date": str(w.end_date),
            "version": w.version}


@router.post("/schedule-windows/{window_id}/pause")
async def pause_window(window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
                       db: AsyncSession = Depends(get_db)) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    if w.status != "active":
        raise HTTPException(status_code=400, detail="רק לוחות פעילים ניתנים להשהייה")
    w.status = "paused"
    w.paused_at = datetime.now(timezone.utc)
    db.add(ScheduleWindowLifecycleEvent(
        tenant_id=tenant.id, schedule_window_id=w.id, event_type="pause",
        performed_by=user.id, state_snapshot={"previous_status": "active"},
    ))
    await db.commit()
    return {"id": str(w.id), "status": "paused"}


@router.post("/schedule-windows/{window_id}/resume")
async def resume_window(window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
                        db: AsyncSession = Depends(get_db)) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    if w.status != "paused":
        raise HTTPException(status_code=400, detail="רק לוחות מושהים ניתנים לחידוש")
    w.status = "active"
    w.paused_at = None
    db.add(ScheduleWindowLifecycleEvent(
        tenant_id=tenant.id, schedule_window_id=w.id, event_type="resume",
        performed_by=user.id, resume_mode="continue_from_expected",
        state_snapshot={"previous_status": "paused"},
    ))
    await db.commit()
    return {"id": str(w.id), "status": "active"}


class ResetWindowRequest(PydanticBaseModel):
    confirmation: str
    note: str | None = None


@router.post("/schedule-windows/{window_id}/reset")
async def reset_window(
    window_id: UUID,
    data: ResetWindowRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Reset a schedule window: delete draft/proposed missions, keep completed ones."""
    if data.confirmation != "אני מאשר איפוס":
        raise HTTPException(status_code=400, detail='יש להקליד "אני מאשר איפוס" לאישור')

    result = await db.execute(
        select(ScheduleWindow).where(
            ScheduleWindow.id == window_id,
            ScheduleWindow.tenant_id == tenant.id,
        )
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")

    # Delete missions in draft/proposed status
    missions_result = await db.execute(
        select(Mission).where(
            Mission.schedule_window_id == window_id,
            Mission.status.in_(["draft", "proposed"]),
        )
    )
    draft_missions = missions_result.scalars().all()
    deleted_count = 0
    for m in draft_missions:
        # Delete assignments first
        await db.execute(
            select(MissionAssignment).where(MissionAssignment.mission_id == m.id)
        )
        assignments = (await db.execute(
            select(MissionAssignment).where(MissionAssignment.mission_id == m.id)
        )).scalars().all()
        for a in assignments:
            await db.delete(a)
        await db.delete(m)
        deleted_count += 1

    # Reset window status to active
    w.status = "active"
    w.paused_at = None

    # Lifecycle event
    db.add(ScheduleWindowLifecycleEvent(
        tenant_id=tenant.id, schedule_window_id=w.id, event_type="reset",
        performed_by=user.id, note=data.note,
        state_snapshot={"deleted_missions": deleted_count},
    ))

    # Audit log
    db.add(AuditLog(
        tenant_id=tenant.id,
        user_id=user.id,
        action="reset",
        entity_type="schedule_window",
        entity_id=window_id,
        after_state={
            "deleted_missions": deleted_count,
            "note": data.note,
        },
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))

    await db.commit()
    return {
        "id": str(w.id),
        "status": w.status,
        "deleted_missions": deleted_count,
        "note": data.note,
    }


@router.post("/schedule-windows/{window_id}/activate")
async def activate_window(window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
                          db: AsyncSession = Depends(get_db)) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    if w.status != "draft":
        raise HTTPException(status_code=400, detail="רק לוחות בטיוטה ניתנים להפעלה")
    w.status = "active"
    db.add(ScheduleWindowLifecycleEvent(
        tenant_id=tenant.id, schedule_window_id=w.id, event_type="activate",
        performed_by=user.id, state_snapshot={"previous_status": "draft"},
    ))
    await db.commit()
    return {"id": str(w.id), "status": "active"}


@router.post("/schedule-windows/{window_id}/archive")
async def archive_window(window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
                         db: AsyncSession = Depends(get_db)) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    previous_status = w.status
    w.status = "archived"
    db.add(ScheduleWindowLifecycleEvent(
        tenant_id=tenant.id, schedule_window_id=w.id, event_type="archive",
        performed_by=user.id, state_snapshot={"previous_status": previous_status},
    ))
    await db.commit()
    return {"id": str(w.id), "status": "archived"}


@router.post("/schedule-windows/{window_id}/copy", status_code=status.HTTP_201_CREATED)
async def copy_window(window_id: UUID, data: ScheduleWindowCreate,
                      tenant: CurrentTenant, user: CurrentUser,
                      db: AsyncSession = Depends(get_db)) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    source = result.scalar_one_or_none()
    if not source:
        raise HTTPException(status_code=404, detail="לוח מקור לא נמצא")
    new_window = ScheduleWindow(
        tenant_id=tenant.id, name=data.name,
        start_date=data.start_date, end_date=data.end_date,
        settings_override=source.settings_override, template_id=source.id,
    )
    db.add(new_window)
    await db.flush()
    # Copy employees
    emp_result = await db.execute(
        select(ScheduleWindowEmployee).where(ScheduleWindowEmployee.schedule_window_id == source.id)
    )
    for swe in emp_result.scalars().all():
        db.add(ScheduleWindowEmployee(
            schedule_window_id=new_window.id, employee_id=swe.employee_id,
        ))
    await db.commit()
    await db.refresh(new_window)
    return {"id": str(new_window.id), "name": new_window.name, "status": new_window.status}


@router.post("/schedule-windows/{window_id}/employees")
async def add_employees_to_window(
    window_id: UUID, data: ScheduleWindowEmployeeAdd,
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    added = 0
    for eid in data.employee_ids:
        existing = await db.execute(
            select(ScheduleWindowEmployee).where(
                ScheduleWindowEmployee.schedule_window_id == window_id,
                ScheduleWindowEmployee.employee_id == eid,
            )
        )
        if not existing.scalar_one_or_none():
            db.add(ScheduleWindowEmployee(schedule_window_id=window_id, employee_id=eid))
            added += 1
    await db.commit()
    return {"added": added}


@router.delete("/schedule-windows/{window_id}/employees/{employee_id}", status_code=204)
async def remove_employee_from_window(
    window_id: UUID, employee_id: UUID,
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
) -> None:
    result = await db.execute(
        select(ScheduleWindowEmployee).where(
            ScheduleWindowEmployee.schedule_window_id == window_id,
            ScheduleWindowEmployee.employee_id == employee_id,
        )
    )
    swe = result.scalar_one_or_none()
    if swe:
        await db.delete(swe)
        await db.commit()


@router.get("/schedule-windows/{window_id}/employees")
async def list_window_employees(
    window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    result = await db.execute(
        select(ScheduleWindowEmployee, Employee)
        .join(Employee, ScheduleWindowEmployee.employee_id == Employee.id)
        .where(ScheduleWindowEmployee.schedule_window_id == window_id)
        .order_by(Employee.full_name)
    )
    rows = result.all()

    # Load work roles for all employees in this window
    emp_ids = [e.id for _, e in rows]
    work_role_map: dict[str, list[dict]] = {}
    if emp_ids:
        from app.models.employee import EmployeeWorkRole
        from app.models.resource import WorkRole
        wr_result = await db.execute(
            select(EmployeeWorkRole, WorkRole)
            .join(WorkRole, EmployeeWorkRole.work_role_id == WorkRole.id)
            .where(EmployeeWorkRole.employee_id.in_(emp_ids))
        )
        for ewr, wr in wr_result.all():
            eid = str(ewr.employee_id)
            if eid not in work_role_map:
                work_role_map[eid] = []
            work_role_map[eid].append({
                "id": str(wr.id),
                "name": wr.name,
                "color": wr.color,
                "is_primary": ewr.is_primary,
            })

    return [
        {
            "id": str(swe.id),
            "employee_id": str(e.id),
            "full_name": e.full_name,
            "employee_number": e.employee_number,
            "status": e.status,
            "is_active": e.is_active,
            "work_roles": work_role_map.get(str(e.id), []),
        }
        for swe, e in rows
    ]


# ═══════════════════════════════════════════
# Mission Types
# ═══════════════════════════════════════════

@router.get("/mission-types")
async def list_mission_types(
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
    active_only: bool = True,
) -> list[dict]:
    query = select(MissionType).where(MissionType.tenant_id == tenant.id)
    if active_only:
        query = query.where(MissionType.is_active.is_(True))
    query = query.order_by(MissionType.created_at)
    result = await db.execute(query)
    return [MissionTypeResponse.model_validate(mt).model_dump() for mt in result.scalars().all()]


@router.post("/mission-types", status_code=status.HTTP_201_CREATED)
async def create_mission_type(
    data: MissionTypeCreate, tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    _validate_jsonb(data.required_slots, MissionSlot, "required_slots")
    if isinstance(data.timeline_items, list):
        _validate_jsonb(data.timeline_items, TimelineItem, "timeline_items")
    mt = MissionType(tenant_id=tenant.id, **data.model_dump())
    db.add(mt)
    await db.flush()
    await db.refresh(mt)
    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="create",
        entity_type="mission_type", entity_id=mt.id,
        after_state={"name": mt.name},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    return MissionTypeResponse.model_validate(mt).model_dump()


@router.get("/mission-types/{mt_id}")
async def get_mission_type(
    mt_id: UUID, tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(MissionType).where(MissionType.id == mt_id, MissionType.tenant_id == tenant.id)
    )
    mt = result.scalar_one_or_none()
    if not mt:
        raise HTTPException(status_code=404, detail="סוג משימה לא נמצא")
    return MissionTypeResponse.model_validate(mt).model_dump()


@router.patch("/mission-types/{mt_id}")
async def update_mission_type(
    mt_id: UUID, data: MissionTypeUpdate, tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(MissionType).where(MissionType.id == mt_id, MissionType.tenant_id == tenant.id)
    )
    mt = result.scalar_one_or_none()
    if not mt:
        raise HTTPException(status_code=404, detail="סוג משימה לא נמצא")
    update_data = data.model_dump(exclude_unset=True)
    if "required_slots" in update_data:
        _validate_jsonb(update_data["required_slots"], MissionSlot, "required_slots")
    if "timeline_items" in update_data and isinstance(update_data["timeline_items"], list):
        _validate_jsonb(update_data["timeline_items"], TimelineItem, "timeline_items")
    for key, value in update_data.items():
        setattr(mt, key, value)
    await db.flush()
    await db.refresh(mt)
    await db.commit()
    return MissionTypeResponse.model_validate(mt).model_dump()


@router.delete("/mission-types/{mt_id}", status_code=204)
async def delete_mission_type(
    mt_id: UUID, tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
) -> None:
    result = await db.execute(
        select(MissionType).where(MissionType.id == mt_id, MissionType.tenant_id == tenant.id)
    )
    mt = result.scalar_one_or_none()
    if not mt:
        raise HTTPException(status_code=404, detail="סוג משימה לא נמצא")

    # Check if active missions use this type
    active_count = (await db.execute(
        select(func.count()).where(
            Mission.mission_type_id == mt_id,
            Mission.status.not_in(["cancelled", "archived"]),
        )
    )).scalar() or 0
    if active_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"לא ניתן למחוק סוג משימה עם {active_count} משימות פעילות"
        )

    mt.is_active = False
    await db.commit()


# ═══════════════════════════════════════════
# Mission Templates
# ═══════════════════════════════════════════

@router.get("/mission-templates")
async def list_mission_templates(
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
    window_id: UUID | None = None,
) -> list[dict]:
    query = select(MissionTemplate).where(MissionTemplate.tenant_id == tenant.id)
    if window_id:
        query = query.where(MissionTemplate.schedule_window_id == window_id)
    result = await db.execute(query.order_by(MissionTemplate.name))
    return [MissionTemplateResponse.model_validate(t).model_dump() for t in result.scalars().all()]


@router.post("/mission-templates", status_code=status.HTTP_201_CREATED)
async def create_mission_template(
    data: MissionTemplateCreate, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    _validate_recurrence(data.recurrence)
    tmpl = MissionTemplate(tenant_id=tenant.id, **data.model_dump())
    db.add(tmpl)
    await db.flush()
    await db.refresh(tmpl)
    await db.commit()
    return MissionTemplateResponse.model_validate(tmpl).model_dump()


@router.patch("/mission-templates/{tmpl_id}")
async def update_mission_template(
    tmpl_id: UUID, data: MissionTemplateUpdate, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(MissionTemplate).where(MissionTemplate.id == tmpl_id, MissionTemplate.tenant_id == tenant.id)
    )
    tmpl = result.scalar_one_or_none()
    if not tmpl:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")
    update_data = data.model_dump(exclude_unset=True)
    if "recurrence" in update_data:
        _validate_recurrence(update_data["recurrence"])
    for key, value in update_data.items():
        setattr(tmpl, key, value)
    await db.flush()
    await db.refresh(tmpl)
    await db.commit()
    return MissionTemplateResponse.model_validate(tmpl).model_dump()


@router.delete("/mission-templates/{tmpl_id}", status_code=204)
async def delete_mission_template(
    tmpl_id: UUID, tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
) -> None:
    result = await db.execute(
        select(MissionTemplate).where(MissionTemplate.id == tmpl_id, MissionTemplate.tenant_id == tenant.id)
    )
    tmpl = result.scalar_one_or_none()
    if not tmpl:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")
    tmpl.is_active = False
    await db.commit()


# ═══════════════════════════════════════════
# Missions
# ═══════════════════════════════════════════

@router.get("/missions")
async def list_missions(
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
    window_id: UUID | None = None, date_from: date | None = None, date_to: date | None = None,
    status_filter: str | None = None, page: int = 1, page_size: int = 200,
) -> list[dict]:
    """List missions with pagination. Returns list for backward compat (frontend expects array)."""
    base_query = select(Mission).where(Mission.tenant_id == tenant.id)
    if window_id:
        base_query = base_query.where(Mission.schedule_window_id == window_id)
    if date_from:
        base_query = base_query.where(Mission.date >= date_from)
    if date_to:
        base_query = base_query.where(Mission.date <= date_to)
    if status_filter:
        base_query = base_query.where(Mission.status == status_filter)

    # Paginate
    query = base_query.order_by(Mission.date, Mission.start_time).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    missions = result.scalars().all()

    # Batch-load mission types to avoid N+1
    mt_ids = list({m.mission_type_id for m in missions})
    mt_map = {}
    if mt_ids:
        mt_result = await db.execute(select(MissionType).where(MissionType.id.in_(mt_ids)))
        for mt in mt_result.scalars().all():
            mt_map[str(mt.id)] = mt

    # Batch-load assignments
    mission_ids = [m.id for m in missions]
    assignment_map: dict[str, list] = {str(mid): [] for mid in mission_ids}
    if mission_ids:
        assign_result = await db.execute(
            select(MissionAssignment, Employee)
            .join(Employee, MissionAssignment.employee_id == Employee.id)
            .where(MissionAssignment.mission_id.in_(mission_ids))
        )
        for ma, emp in assign_result.all():
            assignment_map[str(ma.mission_id)].append({
                "id": str(ma.id),
                "employee_id": str(ma.employee_id),
                "employee_name": emp.full_name,
                "work_role_id": str(ma.work_role_id),
                "slot_id": ma.slot_id,
                "status": ma.status,
                "conflicts_detected": ma.conflicts_detected,
            })

    items = []
    for m in missions:
        mt = mt_map.get(str(m.mission_type_id))
        items.append({
            "id": str(m.id),
            "tenant_id": str(m.tenant_id),
            "schedule_window_id": str(m.schedule_window_id),
            "mission_type_id": str(m.mission_type_id),
            "mission_type_name": mt.name if mt else None,
            "template_id": str(m.template_id) if m.template_id else None,
            "name": m.name,
            "date": str(m.date),
            "start_time": str(m.start_time),
            "end_time": str(m.end_time),
            "status": m.status,
            "is_activated": m.is_activated,
            "version": m.version,
            "assignments": assignment_map.get(str(m.id), []),
            "created_at": str(m.created_at),
            "updated_at": str(m.updated_at),
            # Include required_slots (from mission itself or fall back to mission type)
            "required_slots": m.required_slots or (mt.required_slots if mt else None),
        })
    return items




@router.get("/missions/{mission_id}")
async def get_mission(
    mission_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    """Get a single mission by ID."""
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    mission = result.scalar_one_or_none()
    if not mission:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")

    # Get mission type name
    mt_name = None
    if mission.mission_type_id:
        mt_result = await db.execute(select(MissionType).where(MissionType.id == mission.mission_type_id))
        mt = mt_result.scalar_one_or_none()
        if mt:
            mt_name = mt.name

    # Get assignments
    assign_result = await db.execute(
        select(MissionAssignment, Employee)
        .join(Employee, MissionAssignment.employee_id == Employee.id)
        .where(MissionAssignment.mission_id == mission.id)
    )
    assignments = [
        {
            "id": str(ma.id),
            "employee_id": str(ma.employee_id),
            "employee_name": emp.full_name,
            "work_role_id": str(ma.work_role_id),
            "slot_id": ma.slot_id,
            "status": ma.status,
            "conflicts_detected": ma.conflicts_detected,
        }
        for ma, emp in assign_result.all()
    ]

    return {
        "id": str(mission.id),
        "tenant_id": str(mission.tenant_id),
        "schedule_window_id": str(mission.schedule_window_id),
        "mission_type_id": str(mission.mission_type_id),
        "mission_type_name": mt_name,
        "template_id": str(mission.template_id) if mission.template_id else None,
        "name": mission.name,
        "date": str(mission.date),
        "start_time": str(mission.start_time) if mission.start_time else None,
        "end_time": str(mission.end_time) if mission.end_time else None,
        "status": mission.status,
        "is_activated": mission.is_activated,
        "version": mission.version,
        "notes": getattr(mission, "notes", None),
        "assignments": assignments,
        "created_at": str(mission.created_at),
        "updated_at": str(mission.updated_at),
    }

@router.post("/missions", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permission("missions", "write"))])
async def create_mission(
    data: MissionCreate, tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    # Validate window exists
    w_result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == data.schedule_window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    if not w_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")
    # Validate mission type exists and load its slots
    mt_result = await db.execute(
        select(MissionType).where(MissionType.id == data.mission_type_id, MissionType.tenant_id == tenant.id)
    )
    mt = mt_result.scalar_one_or_none()
    if not mt:
        raise HTTPException(status_code=404, detail="סוג משימה לא נמצא")

    # Copy required_slots from mission type if not explicitly provided
    mission_data = data.model_dump()
    if not mission_data.get("required_slots") and mt.required_slots:
        mission_data["required_slots"] = mt.required_slots

    mission = Mission(
        tenant_id=tenant.id, created_by=user.id,
        **mission_data,
    )
    db.add(mission)
    await db.flush()
    await db.refresh(mission)

    # === Conflict detection: check time overlaps with other missions on same window/date ===
    warnings: list[str] = []
    if mission.date and mission.start_time and mission.end_time:
        overlap_result = await db.execute(
            select(Mission).where(
                Mission.schedule_window_id == mission.schedule_window_id,
                Mission.date == mission.date,
                Mission.id != mission.id,
                Mission.status.not_in(["cancelled", "archived"]),
                Mission.start_time < mission.end_time,
                Mission.end_time > mission.start_time,
            )
        )
        for overlapping in overlap_result.scalars().all():
            warnings.append(f"חפיפת זמנים עם משימה {overlapping.name}")

    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="create",
        entity_type="mission", entity_id=mission.id,
        after_state={"name": mission.name, "date": str(mission.date)},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    result = {
        "id": str(mission.id), "name": mission.name, "date": str(mission.date),
        "start_time": str(mission.start_time), "end_time": str(mission.end_time),
        "status": mission.status, "assignments": [],
        "schedule_window_id": str(mission.schedule_window_id),
        "mission_type_id": str(mission.mission_type_id),
        "required_slots": mission.required_slots,
    }
    if warnings:
        result["warnings"] = warnings
    return result


@router.patch("/missions/{mission_id}")
async def update_mission(
    mission_id: UUID, data: MissionUpdate, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")

    # Concurrent editing protection — require version match
    update_data = data.model_dump(exclude_unset=True)
    client_version = update_data.pop("version", None)
    if client_version is not None and client_version != m.version:
        raise HTTPException(
            status_code=409,
            detail="מישהו אחר ערך את המשימה. רענן ונסה שוב",
        )

    for key, value in update_data.items():
        setattr(m, key, value)
    m.version += 1
    await db.flush()
    await db.refresh(m)
    await db.commit()
    return {"id": str(m.id), "name": m.name, "status": m.status, "version": m.version}


@router.post("/missions/{mission_id}/approve", dependencies=[Depends(require_permission("missions", "approve"))])
async def approve_mission(
    mission_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    m.status = "approved"
    m.approved_by = user.id
    m.approved_at = datetime.now(timezone.utc)
    await db.commit()

    # Notify all assigned employees
    try:
        from app.tasks.notifications import send_notification
        assignments = await db.execute(
            select(MissionAssignment).where(
                MissionAssignment.mission_id == m.id,
                MissionAssignment.status != "replaced",
            )
        )
        for a in assignments.scalars().all():
            send_notification.delay(
                str(tenant.id), str(a.employee_id), "mission_assigned",
                {"mission.name": m.name, "mission.date": str(m.date),
                 "mission.start_time": str(m.start_time)}
            )
    except Exception:
        pass

    # Broadcast mission.approved via WebSocket
    await ws_manager.broadcast_to_tenant(tenant.slug, "mission.approved", {
        "mission_id": str(m.id),
        "name": m.name,
        "approved_by": str(user.id),
    })

    return {"id": str(m.id), "status": "approved"}


@router.post("/missions/{mission_id}/cancel")
async def cancel_mission(
    mission_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    m.status = "cancelled"
    await db.commit()

    # Notify all assigned employees about cancellation
    try:
        from app.tasks.notifications import send_notification
        assignments = await db.execute(
            select(MissionAssignment).where(
                MissionAssignment.mission_id == m.id,
                MissionAssignment.status != "replaced",
            )
        )
        for a in assignments.scalars().all():
            send_notification.delay(
                str(tenant.id), str(a.employee_id), "mission_cancelled",
                {"mission.name": m.name, "mission.date": str(m.date)}
            )
    except Exception:
        pass

    return {"id": str(m.id), "status": "cancelled"}


@router.post("/missions/generate")
async def generate_missions(
    data: MissionGenerateRequest, tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    """Generate mission instances from a template for a date range."""
    tmpl_result = await db.execute(
        select(MissionTemplate).where(MissionTemplate.id == data.template_id, MissionTemplate.tenant_id == tenant.id)
    )
    tmpl = tmpl_result.scalar_one_or_none()
    if not tmpl:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")

    mt_result = await db.execute(select(MissionType).where(MissionType.id == tmpl.mission_type_id))
    mt = mt_result.scalar_one_or_none()

    recurrence = tmpl.recurrence or {"type": "daily"}
    time_slots = tmpl.time_slots or [{"start": "08:00", "end": "16:00"}]

    # === Pre-load existing missions for duplicate detection ===
    existing_missions_result = await db.execute(
        select(Mission).where(
            Mission.template_id == tmpl.id,
            Mission.schedule_window_id == tmpl.schedule_window_id,
            Mission.date >= data.start_date,
            Mission.date <= data.end_date,
            Mission.status.not_in(["cancelled"]),
        )
    )
    existing_by_date_time: set[tuple] = set()
    for em in existing_missions_result.scalars().all():
        existing_by_date_time.add((str(em.date), str(em.start_time), str(em.end_time)))

    created_missions = []
    skipped_duplicates = 0
    current = data.start_date
    while current <= data.end_date:
        should_create = False
        rec_type = recurrence.get("type", "daily")
        if rec_type == "daily":
            should_create = True
        elif rec_type in ("weekly", "custom", "specific_days"):
            days = recurrence.get("days_of_week") or recurrence.get("days", [])
            if current.weekday() in days:
                should_create = True
            # Check active_weeks (odd/even)
            active_weeks = recurrence.get("active_weeks", "all")
            if active_weeks and active_weeks != "all":
                week_num = current.isocalendar()[1]
                if active_weeks == "odd" and week_num % 2 == 0:
                    should_create = False
                elif active_weeks == "even" and week_num % 2 == 1:
                    should_create = False
        elif rec_type == "one_time":
            # Only create on the start_date
            if current == data.start_date:
                should_create = True

        # Check exceptions
        exceptions = recurrence.get("exceptions", [])
        if str(current) in exceptions:
            should_create = False
        # Check extra_dates
        extra_dates = recurrence.get("extra_dates", [])
        if str(current) in extra_dates and not should_create:
            should_create = True

        if should_create:
            for slot in time_slots:
                start_parts = slot.get("start", "08:00").split(":")
                end_parts = slot.get("end", "16:00").split(":")
                # Check for duplicate: same template, date, start/end time
                slot_start = time(int(start_parts[0]), int(start_parts[1]))
                slot_end = time(int(end_parts[0]), int(end_parts[1]))
                dup_key = (str(current), str(slot_start), str(slot_end))
                if dup_key in existing_by_date_time:
                    skipped_duplicates += 1
                    continue
                mission = Mission(
                    tenant_id=tenant.id,
                    schedule_window_id=tmpl.schedule_window_id,
                    mission_type_id=tmpl.mission_type_id,
                    template_id=tmpl.id,
                    name=f"{tmpl.name} - {current.isoformat()}",
                    date=current,
                    start_time=time(int(start_parts[0]), int(start_parts[1])),
                    end_time=time(int(end_parts[0]), int(end_parts[1])),
                    created_by=user.id,
                    # Copy slots from mission type
                    required_slots=mt.required_slots if mt and mt.required_slots else None,
                )
                db.add(mission)
                await db.flush()
                await db.refresh(mission)
                created_missions.append({
                    "id": str(mission.id), "name": mission.name,
                    "date": str(mission.date), "status": mission.status,
                    "parent_mission_id": None,
                })

                # ═══ POST-MISSION RULE: Auto-create follow-up missions ═══
                # E.g., patrol → standby: auto-creates standby mission linked to this patrol
                post_rule = mt.post_mission_rule if mt else None
                if post_rule and post_rule.get("auto_transition_to_mission_type_id"):
                    followup_mt_id = post_rule["auto_transition_to_mission_type_id"]
                    followup_mt_result = await db.execute(
                        select(MissionType).where(MissionType.id == followup_mt_id)
                    )
                    followup_mt = followup_mt_result.scalar_one_or_none()

                    if followup_mt:
                        # Follow-up starts when parent ends
                        followup_start = mission.end_time
                        # Duration from the follow-up mission type, default 8h
                        followup_duration = followup_mt.duration_hours or 8
                        followup_end_hour = (followup_start.hour + int(followup_duration)) % 24
                        followup_end_min = followup_start.minute
                        # Date: if crosses midnight, next day
                        followup_date = current
                        if followup_start.hour + int(followup_duration) >= 24:
                            followup_date = current + timedelta(days=1)

                        followup_name_he = followup_mt.name.get("he", "") if isinstance(followup_mt.name, dict) else str(followup_mt.name)
                        parent_name = tmpl.name

                        followup_mission = Mission(
                            tenant_id=tenant.id,
                            schedule_window_id=tmpl.schedule_window_id,
                            mission_type_id=followup_mt_id,
                            template_id=tmpl.id,
                            name=f"{followup_name_he} (אחרי {parent_name}) - {current.isoformat()}",
                            date=followup_date,
                            start_time=followup_start,
                            end_time=time(followup_end_hour, followup_end_min),
                            created_by=user.id,
                            parent_mission_id=mission.id,
                            post_mission_config={
                                "source_rule": post_rule,
                                "auto_assign_same_crew": post_rule.get("auto_assign_same_crew", True),
                                "condition": post_rule.get("condition", "always"),
                            },
                        )
                        db.add(followup_mission)
                        await db.flush()
                        await db.refresh(followup_mission)

                        # Auto-assign same crew if configured
                        if post_rule.get("auto_assign_same_crew", True):
                            # Copy assignments from parent mission to follow-up
                            parent_assignments = await db.execute(
                                select(MissionAssignment).where(
                                    MissionAssignment.mission_id == mission.id,
                                    MissionAssignment.status != "replaced",
                                )
                            )
                            for pa in parent_assignments.scalars().all():
                                # Try to match slot — if follow-up has matching role, use it
                                followup_slots = followup_mt.required_slots or []
                                target_slot = None
                                for fs in followup_slots:
                                    if fs.get("work_role_id") == str(pa.work_role_id):
                                        target_slot = fs.get("slot_id", "default")
                                        break
                                if not target_slot and followup_slots:
                                    target_slot = followup_slots[0].get("slot_id", "default")

                                if target_slot:
                                    followup_assignment = MissionAssignment(
                                        mission_id=followup_mission.id,
                                        employee_id=pa.employee_id,
                                        work_role_id=pa.work_role_id,
                                        slot_id=target_slot,
                                        status="assigned",
                                    )
                                    db.add(followup_assignment)

                        created_missions.append({
                            "id": str(followup_mission.id),
                            "name": followup_mission.name,
                            "date": str(followup_mission.date),
                            "status": followup_mission.status,
                            "parent_mission_id": str(mission.id),
                            "auto_created": True,
                        })

        current += timedelta(days=1)

    await db.commit()
    return {
        "created": len(created_missions),
        "skipped_duplicates": skipped_duplicates,
        "missions": created_missions,
        "follow_ups_created": sum(1 for m in created_missions if m.get("auto_created")),
    }


# ═══════════════════════════════════════════
# Mission Assignments
# ═══════════════════════════════════════════

@router.get("/missions/{mission_id}/assignments")
async def list_assignments(
    mission_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    result = await db.execute(
        select(MissionAssignment, Employee)
        .join(Employee, MissionAssignment.employee_id == Employee.id)
        .where(MissionAssignment.mission_id == mission_id)
    )
    return [
        {
            "id": str(ma.id),
            "mission_id": str(ma.mission_id),
            "employee_id": str(ma.employee_id),
            "employee_name": emp.full_name,
            "work_role_id": str(ma.work_role_id),
            "slot_id": ma.slot_id,
            "status": ma.status,
            "conflicts_detected": ma.conflicts_detected,
            "assigned_at": str(ma.assigned_at) if ma.assigned_at else None,
        }
        for ma, emp in result.all()
    ]


@router.post("/missions/{mission_id}/assignments", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permission("missions", "write"))])
async def create_assignment(
    mission_id: UUID, data: MissionAssignmentCreate,
    tenant: CurrentTenant, user: CurrentUser, request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    # Validate mission
    m_result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    mission = m_result.scalar_one_or_none()
    if not mission:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")

    # Check: same employee already in this mission?
    dup_check = await db.execute(
        select(MissionAssignment).where(
            MissionAssignment.mission_id == mission_id,
            MissionAssignment.employee_id == data.employee_id,
            MissionAssignment.status != "replaced",
        )
    )
    if dup_check.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="החייל כבר משובץ למשימה זו",
        )

    # Check for conflicts (same employee, overlapping time on same date)
    conflicts = []
    conflict_result = await db.execute(
        select(MissionAssignment, Mission)
        .join(Mission, MissionAssignment.mission_id == Mission.id)
        .where(
            MissionAssignment.employee_id == data.employee_id,
            Mission.date == mission.date,
            Mission.id != mission_id,
            MissionAssignment.status != "replaced",
        )
    )
    for existing_ma, existing_m in conflict_result.all():
        # Simple time overlap check
        if (existing_m.start_time < mission.end_time and existing_m.end_time > mission.start_time):
            conflicts.append({
                "type": "time_overlap",
                "mission_id": str(existing_m.id),
                "mission_name": existing_m.name,
                "time": f"{existing_m.start_time}-{existing_m.end_time}",
            })

    assignment = MissionAssignment(
        mission_id=mission_id,
        employee_id=data.employee_id,
        work_role_id=data.work_role_id,
        slot_id=data.slot_id,
        assigned_at=datetime.now(timezone.utc),
        conflicts_detected=conflicts if conflicts else None,
    )
    db.add(assignment)
    await db.flush()
    await db.refresh(assignment)

    # Get employee name
    emp_result = await db.execute(select(Employee).where(Employee.id == data.employee_id))
    emp = emp_result.scalar_one_or_none()

    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="assign",
        entity_type="mission_assignment", entity_id=assignment.id,
        after_state={"employee": emp.full_name if emp else str(data.employee_id), "mission": mission.name},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()

    # Send notification to assigned employee
    try:
        from app.tasks.notifications import send_notification
        send_notification.delay(
            str(tenant.id), str(data.employee_id), "mission_assigned",
            {"mission.name": mission.name, "mission.date": str(mission.date),
             "mission.start_time": str(mission.start_time), "employee.name": emp.full_name if emp else ""}
        )
    except Exception:
        pass  # Don't fail assignment if notification fails

    return {
        "id": str(assignment.id),
        "mission_id": str(mission_id),
        "employee_id": str(data.employee_id),
        "employee_name": emp.full_name if emp else None,
        "work_role_id": str(data.work_role_id),
        "slot_id": data.slot_id,
        "status": assignment.status,
        "conflicts_detected": assignment.conflicts_detected,
        "assigned_at": str(assignment.assigned_at),
    }


@router.delete("/missions/{mission_id}/assignments/{assignment_id}", status_code=204)
async def remove_assignment(
    mission_id: UUID, assignment_id: UUID,
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
) -> None:
    result = await db.execute(
        select(MissionAssignment).where(
            MissionAssignment.id == assignment_id,
            MissionAssignment.mission_id == mission_id,
        )
    )
    ma = result.scalar_one_or_none()
    if not ma:
        raise HTTPException(status_code=404, detail="שיבוץ לא נמצא")
    ma.status = "replaced"
    await db.commit()


# ═══════════════════════════════════════════
# Smart Manual Assignment — Eligible Soldiers
# ═══════════════════════════════════════════

@router.get("/missions/{mission_id}/eligible-soldiers/{slot_id}")
async def get_eligible_soldiers(
    mission_id: UUID, slot_id: str,
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """Get eligible soldiers for a mission slot, scored and sorted."""
    # Get mission
    m_result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    mission = m_result.scalar_one_or_none()
    if not mission:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")

    # Get mission type to find slot's required work_role
    mt_result = await db.execute(select(MissionType).where(MissionType.id == mission.mission_type_id))
    mt = mt_result.scalar_one_or_none()
    required_work_role_id = None
    if mt and mt.required_slots:
        for slot in mt.required_slots:
            if slot.get("slot_id") == slot_id:
                required_work_role_id = slot.get("work_role_id")
                break

    # Get employees in the schedule window
    emp_query = (
        select(Employee)
        .join(ScheduleWindowEmployee, ScheduleWindowEmployee.employee_id == Employee.id)
        .where(
            ScheduleWindowEmployee.schedule_window_id == mission.schedule_window_id,
            Employee.is_active.is_(True),
        )
    )
    emp_result = await db.execute(emp_query)
    employees = emp_result.scalars().all()

    # Fallback: if no employees in the schedule window, use all active employees in tenant
    if not employees:
        fallback_result = await db.execute(
            select(Employee).where(
                Employee.tenant_id == tenant.id,
                Employee.is_active.is_(True),
            )
        )
        employees = fallback_result.scalars().all()

    # If required_work_role_id, filter by work role
    if required_work_role_id:
        wr_result = await db.execute(
            select(EmployeeWorkRole.employee_id).where(
                EmployeeWorkRole.work_role_id == required_work_role_id
            )
        )
        valid_emp_ids = {row[0] for row in wr_result.all()}
        employees = [e for e in employees if e.id in valid_emp_ids]

    # Include all employees — if status is not "present", add a warning but don't exclude
    results = []
    for emp in employees:
        warnings = []
        score = 100

        # Warn if not present
        if emp.status not in ("present", "returning_home", "training"):
            score -= 40
            warnings.append(f"סטטוס נוכחות: {emp.status}")

        # Check time conflicts on same day
        conflict_result = await db.execute(
            select(MissionAssignment, Mission)
            .join(Mission, MissionAssignment.mission_id == Mission.id)
            .where(
                MissionAssignment.employee_id == emp.id,
                Mission.date == mission.date,
                MissionAssignment.status != "replaced",
            )
        )
        day_assignments = conflict_result.all()
        has_hard_conflict = False

        for ma, existing_m in day_assignments:
            # Check time overlap
            if existing_m.start_time < mission.end_time and existing_m.end_time > mission.start_time:
                has_hard_conflict = True
                break
            score -= 10  # Penalty per existing assignment

        if has_hard_conflict:
            continue  # Skip soldiers with hard time conflicts

        # Check rest hours since last mission (look at yesterday + today)
        yesterday = mission.date - timedelta(days=1)
        recent_result = await db.execute(
            select(Mission)
            .join(MissionAssignment, MissionAssignment.mission_id == Mission.id)
            .where(
                MissionAssignment.employee_id == emp.id,
                Mission.date.in_([yesterday, mission.date]),
                MissionAssignment.status != "replaced",
            )
            .order_by(Mission.date.desc(), Mission.end_time.desc())
        )
        recent_missions = recent_result.scalars().all()
        if recent_missions:
            last = recent_missions[0]
            # Calculate hours since last mission ended
            last_end = datetime.combine(last.date, last.end_time)
            this_start = datetime.combine(mission.date, mission.start_time)
            hours_rest = (this_start - last_end).total_seconds() / 3600
            if hours_rest < 16:
                score -= 30
                warnings.append(f"{hours_rest:.0f} שעות מנוחה בלבד (מינימום 16)")
            elif hours_rest < 18:
                score -= 20
                warnings.append(f"{hours_rest:.0f} שעות מנוחה בלבד (מינימום 18)")

        # Already assigned today count
        if len(day_assignments) > 0:
            warnings.append(f"כבר משובץ ל-{len(day_assignments)} משימות היום")

        is_recommended = score >= 70 and len(warnings) == 0

        results.append({
            "employee_id": str(emp.id),
            "employee_name": emp.full_name,
            "employee_number": emp.employee_number,
            "score": max(0, score),
            "warnings": warnings,
            "is_recommended": is_recommended,
            "status": emp.status,
        })

    # Sort by score descending
    results.sort(key=lambda x: x["score"], reverse=True)
    return results


# ═══════════════════════════════════════════
# Auto-Scheduling
# ═══════════════════════════════════════════

class AutoAssignRequest(PydanticBaseModel):
    schedule_window_id: UUID | None = None
    window_id: UUID | None = None
    date_from: date | None = None
    date_to: date | None = None


@router.post("/missions/auto-assign", status_code=status.HTTP_200_OK, dependencies=[Depends(require_permission("missions", "write"))])
async def auto_assign_missions(
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    body: AutoAssignRequest | None = None,
    window_id: UUID | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> dict:
    """Production auto-scheduling: hard filter → scoring → preference optimization → conflict check."""
    from app.services.scheduling_service import AutoScheduler

    # Accept window_id from query param or body (schedule_window_id or window_id)
    resolved_window_id = window_id
    if not resolved_window_id and body:
        resolved_window_id = body.schedule_window_id or body.window_id
    if body and not date_from:
        date_from = body.date_from
    if body and not date_to:
        date_to = body.date_to

    if not resolved_window_id:
        raise HTTPException(status_code=400, detail="נדרש מזהה לוח עבודה")
    window_id = resolved_window_id

    scheduler = AutoScheduler(db, tenant.id, user.id)
    result = await scheduler.run(window_id, date_from, date_to)

    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])

    return result


# Legacy fallback kept for backwards compat
@router.post("/missions/auto-assign-simple", status_code=status.HTTP_200_OK)
async def auto_assign_missions_simple(
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    window_id: UUID | None = None,
) -> dict:
    """Simple auto-assignment fallback."""
    query = select(Mission).where(
        Mission.tenant_id == tenant.id,
        Mission.status == "draft",
    )
    if window_id:
        query = query.where(Mission.schedule_window_id == window_id)
    missions_result = await db.execute(query.order_by(Mission.date, Mission.start_time))
    missions = missions_result.scalars().all()

    total_assigned = 0
    for mission in missions:
        mt_result = await db.execute(select(MissionType).where(MissionType.id == mission.mission_type_id))
        mt = mt_result.scalar_one_or_none()
        if not mt or not mt.required_slots:
            continue

        required_slots = mt.required_slots if isinstance(mt.required_slots, list) else []
        existing = await db.execute(
            select(MissionAssignment).where(
                MissionAssignment.mission_id == mission.id,
                MissionAssignment.status != "replaced",
            )
        )
        filled_slots = {ma.slot_id for ma in existing.scalars().all()}

        window_emps = await db.execute(
            select(ScheduleWindowEmployee, Employee)
            .join(Employee, ScheduleWindowEmployee.employee_id == Employee.id)
            .where(
                ScheduleWindowEmployee.schedule_window_id == mission.schedule_window_id,
                Employee.is_active.is_(True),
            )
        )
        available_employees = [(swe, emp) for swe, emp in window_emps.all()]

        for slot in required_slots:
            slot_id = slot.get("slot_id", "default")
            work_role_id = slot.get("work_role_id")
            count = slot.get("count", 1)

            if slot_id in filled_slots:
                continue

            for i in range(count):
                slot_key = f"{slot_id}_{i}" if count > 1 else slot_id
                if slot_key in filled_slots:
                    continue

                for swe, emp in available_employees:
                    already = await db.execute(
                        select(MissionAssignment).where(
                            MissionAssignment.mission_id == mission.id,
                            MissionAssignment.employee_id == emp.id,
                            MissionAssignment.status != "replaced",
                        )
                    )
                    if already.scalar_one_or_none():
                        continue

                    # Check time conflicts
                    conflict_check = await db.execute(
                        select(MissionAssignment)
                        .join(Mission, MissionAssignment.mission_id == Mission.id)
                        .where(
                            MissionAssignment.employee_id == emp.id,
                            Mission.date == mission.date,
                            Mission.start_time < mission.end_time,
                            Mission.end_time > mission.start_time,
                            MissionAssignment.status != "replaced",
                        )
                    )
                    if conflict_check.scalar_one_or_none():
                        continue

                    # Assign
                    assignment = MissionAssignment(
                        mission_id=mission.id,
                        employee_id=emp.id,
                        work_role_id=work_role_id or emp.id,  # fallback
                        slot_id=slot_key,
                        assigned_at=datetime.now(timezone.utc),
                    )
                    db.add(assignment)
                    total_assigned += 1
                    filled_slots.add(slot_key)
                    break

    await db.commit()
    return {"total_assigned": total_assigned, "missions_processed": len(missions)}


# ═══════════════════════════════════════════
# Swap Requests
# ═══════════════════════════════════════════

@router.get("/swap-requests")
async def list_swap_requests(
    tenant: CurrentTenant, user: CurrentUser, db: AsyncSession = Depends(get_db),
    status_filter: str | None = None,
) -> list[dict]:
    query = select(SwapRequest).where(SwapRequest.tenant_id == tenant.id)
    if status_filter:
        query = query.where(SwapRequest.status == status_filter)
    query = query.order_by(SwapRequest.created_at.desc())
    result = await db.execute(query)
    items = []
    for sr in result.scalars().all():
        # Get requester name
        req_emp = await db.execute(select(Employee).where(Employee.id == sr.requester_employee_id))
        requester = req_emp.scalar_one_or_none()
        target_name = None
        if sr.target_employee_id:
            tgt_emp = await db.execute(select(Employee).where(Employee.id == sr.target_employee_id))
            target = tgt_emp.scalar_one_or_none()
            target_name = target.full_name if target else None
        items.append({
            "id": str(sr.id),
            "requester_employee_id": str(sr.requester_employee_id),
            "requester_name": requester.full_name if requester else None,
            "target_employee_id": str(sr.target_employee_id) if sr.target_employee_id else None,
            "target_name": target_name,
            "swap_type": sr.swap_type,
            "reason": sr.reason,
            "status": sr.status,
            "validation_result": sr.validation_result,
            "target_response": sr.target_response,
            "created_at": str(sr.created_at),
        })
    return items


@router.post("/swap-requests", status_code=status.HTTP_201_CREATED)
async def create_swap_request(
    data: SwapRequestCreate, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    # Validate assignment exists
    assign_result = await db.execute(
        select(MissionAssignment).where(MissionAssignment.id == data.requester_assignment_id)
    )
    assignment = assign_result.scalar_one_or_none()
    if not assignment:
        raise HTTPException(status_code=404, detail="שיבוץ לא נמצא")

    sr = SwapRequest(
        tenant_id=tenant.id,
        requester_employee_id=assignment.employee_id,
        requester_assignment_id=data.requester_assignment_id,
        target_employee_id=data.target_employee_id,
        target_assignment_id=data.target_assignment_id,
        swap_type=data.swap_type,
        reason=data.reason,
    )
    db.add(sr)
    await db.flush()
    await db.refresh(sr)
    await db.commit()
    return {"id": str(sr.id), "status": sr.status, "swap_type": sr.swap_type}


@router.post("/swap-requests/{sr_id}/approve")
async def approve_swap_request(
    sr_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(SwapRequest).where(SwapRequest.id == sr_id, SwapRequest.tenant_id == tenant.id)
    )
    sr = result.scalar_one_or_none()
    if not sr:
        raise HTTPException(status_code=404, detail="בקשת החלפה לא נמצאה")
    if sr.status != "pending":
        raise HTTPException(status_code=400, detail="ניתן לאשר רק בקשות ממתינות")
    sr.status = "approved"
    sr.approved_by = user.id
    await db.commit()

    # Notify requester that swap was approved
    try:
        from app.tasks.notifications import send_notification
        send_notification.delay(str(tenant.id), str(sr.requester_employee_id), "swap_approved", {})
        if sr.target_employee_id:
            send_notification.delay(str(tenant.id), str(sr.target_employee_id), "swap_approved", {})
    except Exception:
        pass

    # Broadcast swap.status_changed via WebSocket
    await ws_manager.broadcast_to_tenant(tenant.slug, "swap.status_changed", {
        "swap_id": str(sr.id),
        "status": "approved",
        "approved_by": str(user.id),
    })

    return {"id": str(sr.id), "status": "approved"}


@router.post("/swap-requests/{sr_id}/reject")
async def reject_swap_request(
    sr_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(SwapRequest).where(SwapRequest.id == sr_id, SwapRequest.tenant_id == tenant.id)
    )
    sr = result.scalar_one_or_none()
    if not sr:
        raise HTTPException(status_code=404, detail="בקשת החלפה לא נמצאה")
    sr.status = "rejected"
    await db.commit()

    # Broadcast swap.status_changed via WebSocket
    await ws_manager.broadcast_to_tenant(tenant.slug, "swap.status_changed", {
        "swap_id": str(sr.id),
        "status": "rejected",
    })

    return {"id": str(sr.id), "status": "rejected"}


# ═══════════════════════════════════════════
# Schedule Window Export/Import Templates
# ═══════════════════════════════════════════

@router.post("/schedule-windows/{window_id}/export-template", dependencies=[Depends(require_permission("missions", "read"))])
async def export_window_template(
    window_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Export window configuration as JSON template (no PII)."""
    result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    w = result.scalar_one_or_none()
    if not w:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")

    # Get mission templates for this window
    tmpl_result = await db.execute(
        select(MissionTemplate).where(
            MissionTemplate.schedule_window_id == window_id,
            MissionTemplate.tenant_id == tenant.id,
        )
    )
    templates = [
        {
            "name": t.name,
            "mission_type_id": str(t.mission_type_id),
            "recurrence": t.recurrence,
            "time_slots": t.time_slots,
            "is_active": t.is_active,
        }
        for t in tmpl_result.scalars().all()
    ]

    # Get mission types used
    mt_ids = list({t["mission_type_id"] for t in templates})
    roles = []
    if mt_ids:
        mt_result = await db.execute(
            select(MissionType).where(MissionType.id.in_([UUID(mid) for mid in mt_ids]))
        )
        for mt in mt_result.scalars().all():
            roles.append({
                "id": str(mt.id),
                "name": mt.name,
                "required_slots": mt.required_slots,
                "color": mt.color,
                "icon": mt.icon,
            })

    # Get rules/settings
    return {
        "export_version": "1.0",
        "name": w.name,
        "start_date": str(w.start_date),
        "end_date": str(w.end_date),
        "settings_override": w.settings_override,
        "notes": w.notes,
        "templates": templates,
        "mission_types": roles,
        "statuses": ["draft", "active", "paused", "archived"],
    }


class ImportTemplateRequest(PydanticBaseModel):
    template: dict
    name: str | None = None


@router.post("/schedule-windows/import-template", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permission("missions", "write"))])
async def import_window_template(
    data: ImportTemplateRequest,
    tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    """Import JSON template → create draft schedule window."""
    tmpl = data.template
    name = data.name or tmpl.get("name", "לוח מיובא")
    start_date_str = tmpl.get("start_date")
    end_date_str = tmpl.get("end_date")

    if not start_date_str or not end_date_str:
        raise HTTPException(status_code=400, detail="תבנית חייבת לכלול תאריכי התחלה וסיום")

    window = ScheduleWindow(
        tenant_id=tenant.id,
        name=name,
        start_date=date.fromisoformat(start_date_str),
        end_date=date.fromisoformat(end_date_str),
        settings_override=tmpl.get("settings_override"),
        notes=tmpl.get("notes"),
        status="draft",
    )
    db.add(window)
    await db.flush()
    await db.refresh(window)

    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="import_template",
        entity_type="schedule_window", entity_id=window.id,
        after_state={"name": window.name, "source": "template_import"},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()

    return {
        "id": str(window.id),
        "name": window.name,
        "status": window.status,
        "start_date": str(window.start_date),
        "end_date": str(window.end_date),
    }


# ═══════════════════════════════════════════
# Mission Activation & Override
# ═══════════════════════════════════════════

@router.post("/missions/{mission_id}/mark-activated", dependencies=[Depends(require_permission("missions", "write"))])
async def mark_mission_activated(
    mission_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    """Mark a standby mission as activated (set is_activated=True)."""
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    if m.is_activated:
        raise HTTPException(status_code=400, detail="משימה כבר מסומנת כמופעלת")

    m.is_activated = True
    m.version += 1

    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="mark_activated",
        entity_type="mission", entity_id=m.id,
        after_state={"is_activated": True, "name": m.name},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    return {"id": str(m.id), "name": m.name, "is_activated": True, "version": m.version}


class OverrideRequest(PydanticBaseModel):
    justification: str


@router.post("/missions/{mission_id}/assignments/{assignment_id}/override", dependencies=[Depends(require_permission("missions", "write"))])
async def override_assignment_conflict(
    mission_id: UUID, assignment_id: UUID,
    data: OverrideRequest,
    tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    """Override a conflict on an assignment with justification text."""
    result = await db.execute(
        select(MissionAssignment).where(
            MissionAssignment.id == assignment_id,
            MissionAssignment.mission_id == mission_id,
        )
    )
    ma = result.scalar_one_or_none()
    if not ma:
        raise HTTPException(status_code=404, detail="שיבוץ לא נמצא")

    if not data.justification.strip():
        raise HTTPException(status_code=400, detail="נדרשת הצדקה לדריסת קונפליקט")

    ma.override_approved_by = user.id
    ma.conflicts_detected = {
        **(ma.conflicts_detected or {}),
        "override_justification": data.justification,
        "override_by": str(user.id),
    }

    # Also update mission override_justification
    m_result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    mission = m_result.scalar_one_or_none()
    if mission:
        mission.override_justification = data.justification

    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="override_conflict",
        entity_type="mission_assignment", entity_id=assignment_id,
        after_state={"justification": data.justification, "mission_id": str(mission_id)},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    return {
        "id": str(ma.id),
        "mission_id": str(mission_id),
        "override_approved_by": str(user.id),
        "justification": data.justification,
    }


# ═══════════════════════════════════════════
# Daily Board Templates
# ═══════════════════════════════════════════

@router.get("/daily-board-templates")
async def list_daily_board_templates(
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.tenant_id == tenant.id,
            DailyBoardTemplate.is_active.is_(True),
        ).order_by(DailyBoardTemplate.created_at)
    )
    return [
        {
            "id": str(t.id),
            "name": t.name,
            "description": t.description,
            "layout": t.layout,
            "columns": t.columns,
            "filters": t.filters,
            "is_default": t.is_default,
            "created_at": str(t.created_at),
            "updated_at": str(t.updated_at),
        }
        for t in result.scalars().all()
    ]


class DailyBoardTemplateCreate(PydanticBaseModel):
    name: str
    description: str | None = None
    layout: dict | None = None
    columns: dict | None = None
    filters: dict | None = None


class DailyBoardTemplateUpdate(PydanticBaseModel):
    name: str | None = None
    description: str | None = None
    layout: dict | None = None
    columns: dict | None = None
    filters: dict | None = None


@router.post("/daily-board-templates", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permission("settings", "write"))])
async def create_daily_board_template(
    data: DailyBoardTemplateCreate,
    tenant: CurrentTenant, user: CurrentUser,
    request: Request, db: AsyncSession = Depends(get_db),
) -> dict:
    tmpl = DailyBoardTemplate(
        tenant_id=tenant.id,
        name=data.name,
        description=data.description,
        layout=data.layout,
        columns=data.columns,
        filters=data.filters,
    )
    db.add(tmpl)
    await db.flush()
    await db.refresh(tmpl)
    db.add(AuditLog(
        tenant_id=tenant.id, user_id=user.id, action="create",
        entity_type="daily_board_template", entity_id=tmpl.id,
        after_state={"name": tmpl.name},
        ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
    ))
    await db.commit()
    return {
        "id": str(tmpl.id),
        "name": tmpl.name,
        "description": tmpl.description,
        "is_default": tmpl.is_default,
        "created_at": str(tmpl.created_at),
    }


@router.get("/daily-board-templates/{template_id}")
async def get_daily_board_template(
    template_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")
    return {
        "id": str(t.id),
        "name": t.name,
        "description": t.description,
        "layout": t.layout,
        "columns": t.columns,
        "filters": t.filters,
        "is_default": t.is_default,
        "created_at": str(t.created_at),
        "updated_at": str(t.updated_at),
    }


@router.patch("/daily-board-templates/{template_id}", dependencies=[Depends(require_permission("settings", "write"))])
async def update_daily_board_template(
    template_id: UUID, data: DailyBoardTemplateUpdate,
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(t, key, value)
    await db.flush()
    await db.refresh(t)
    await db.commit()
    return {
        "id": str(t.id),
        "name": t.name,
        "description": t.description,
        "layout": t.layout,
        "columns": t.columns,
        "filters": t.filters,
        "is_default": t.is_default,
        "updated_at": str(t.updated_at),
    }


@router.delete("/daily-board-templates/{template_id}", status_code=204, dependencies=[Depends(require_permission("settings", "write"))])
async def delete_daily_board_template(
    template_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> None:
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")
    t.is_active = False
    await db.commit()


@router.post("/daily-board-templates/{template_id}/set-default", dependencies=[Depends(require_permission("settings", "write"))])
async def set_default_daily_board_template(
    template_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Set a daily board template as the default (unset others)."""
    # Unset all defaults for this tenant
    all_result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.tenant_id == tenant.id,
            DailyBoardTemplate.is_default.is_(True),
        )
    )
    for t in all_result.scalars().all():
        t.is_default = False

    # Set the new default
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")
    t.is_default = True
    await db.commit()
    return {"id": str(t.id), "name": t.name, "is_default": True}


class BoardPreviewRequest(PydanticBaseModel):
    date: date


@router.post("/daily-board-templates/{template_id}/preview")
async def preview_daily_board_template(
    template_id: UUID, data: BoardPreviewRequest,
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Render a preview of the daily board for a given date."""
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")

    # Fetch missions for the date
    missions_result = await db.execute(
        select(Mission).where(
            Mission.tenant_id == tenant.id,
            Mission.date == data.date,
            Mission.status.not_in(["cancelled"]),
        ).order_by(Mission.start_time)
    )
    missions = missions_result.scalars().all()

    mission_items = []
    for m in missions:
        # Get assignments
        assign_result = await db.execute(
            select(MissionAssignment, Employee)
            .join(Employee, MissionAssignment.employee_id == Employee.id)
            .where(
                MissionAssignment.mission_id == m.id,
                MissionAssignment.status != "replaced",
            )
        )
        assignments = [
            {
                "employee_name": emp.full_name,
                "slot_id": ma.slot_id,
                "status": ma.status,
            }
            for ma, emp in assign_result.all()
        ]

        mt_result = await db.execute(select(MissionType).where(MissionType.id == m.mission_type_id))
        mt = mt_result.scalar_one_or_none()

        mission_items.append({
            "id": str(m.id),
            "name": m.name,
            "mission_type_name": mt.name if mt else None,
            "mission_type_color": mt.color if mt else None,
            "start_time": str(m.start_time),
            "end_time": str(m.end_time),
            "status": m.status,
            "is_activated": m.is_activated,
            "assignments": assignments,
        })

    return {
        "template_id": str(t.id),
        "template_name": t.name,
        "date": str(data.date),
        "layout": t.layout,
        "columns": t.columns,
        "missions": mission_items,
    }


@router.post("/daily-board-templates/{template_id}/export")
async def export_daily_board_template(
    template_id: UUID,
    tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    target_date: date | None = None,
    format: str = "json",
) -> dict:
    """Export daily board template as PDF/Excel placeholder."""
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")

    if format not in ("json", "pdf", "excel"):
        raise HTTPException(status_code=400, detail="פורמט לא נתמך. אפשרויות: json, pdf, excel")

    if format in ("pdf", "excel"):
        # Placeholder — actual generation would use a library
        return {
            "message": f"ייצוא בפורמט {format} בפיתוח",
            "template_id": str(t.id),
            "template_name": t.name,
            "format": format,
        }

    return {
        "template_id": str(t.id),
        "name": t.name,
        "description": t.description,
        "layout": t.layout,
        "columns": t.columns,
        "filters": t.filters,
        "is_default": t.is_default,
        "format": "json",
    }


class BoardGenerateRequest(PydanticBaseModel):
    date_from: date
    date_to: date


@router.post("/daily-board-templates/{template_id}/generate", dependencies=[Depends(require_permission("settings", "write"))])
async def generate_daily_boards_from_template(
    template_id: UUID,
    data: BoardGenerateRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Generate daily boards (missions) from a board template for a date range.

    Uses the template's mission type slots to create missions for each date
    in the range [date_from, date_to].
    """
    # Validate template exists
    result = await db.execute(
        select(DailyBoardTemplate).where(
            DailyBoardTemplate.id == template_id,
            DailyBoardTemplate.tenant_id == tenant.id,
        )
    )
    tmpl = result.scalar_one_or_none()
    if not tmpl:
        raise HTTPException(status_code=404, detail="תבנית לוח יומי לא נמצאה")

    if data.date_from > data.date_to:
        raise HTTPException(status_code=400, detail="תאריך התחלה חייב להיות לפני תאריך סיום")

    max_days = 90
    delta_days = (data.date_to - data.date_from).days + 1
    if delta_days > max_days:
        raise HTTPException(status_code=400, detail=f"מקסימום {max_days} ימים בפעם אחת")

    # Get all mission types for the tenant (for creating missions from template slots)
    mt_result = await db.execute(
        select(MissionType).where(MissionType.tenant_id == tenant.id)
    )
    mission_types = {str(mt.id): mt for mt in mt_result.scalars().all()}

    # We also need a fallback schedule window
    sw_result = await db.execute(
        select(ScheduleWindow).where(
            ScheduleWindow.tenant_id == tenant.id,
            ScheduleWindow.status == "active",
        ).order_by(ScheduleWindow.start_date.desc()).limit(1)
    )
    schedule_window = sw_result.scalar_one_or_none()

    created_count = 0
    skipped_count = 0
    import uuid as _uuid

    current_date = data.date_from
    while current_date <= data.date_to:
        # Check if missions already exist for this date (avoid duplicates)
        existing = await db.execute(
            select(func.count()).select_from(Mission).where(
                Mission.tenant_id == tenant.id,
                Mission.date == current_date,
                Mission.status != "cancelled",
            )
        )
        existing_count = existing.scalar() or 0

        if existing_count > 0:
            skipped_count += 1
            current_date += timedelta(days=1)
            continue

        # Parse template columns to create missions
        # Template columns contain mission type references; layout has the structure
        columns = tmpl.columns or []
        layout = tmpl.layout or {}

        # If the template has structured sections with mission_type_ids, use those
        # Otherwise create a generic "daily board" mission for the date
        sections = layout.get("sections", [])
        missions_created_for_date = False

        for section in sections:
            cells = section.get("cells", [])
            for cell in cells:
                mt_id = cell.get("missionTypeId") or cell.get("mission_type_id")
                if mt_id and mt_id in mission_types:
                    mt = mission_types[mt_id]
                    time_range = cell.get("timeRange") or cell.get("time_range", {})
                    start_t = time_range.get("start", "08:00")
                    end_t = time_range.get("end", "16:00")

                    try:
                        s_parts = start_t.split(":")
                        e_parts = end_t.split(":")
                        start_time = time(int(s_parts[0]), int(s_parts[1]))
                        end_time = time(int(e_parts[0]), int(e_parts[1]))
                    except (ValueError, IndexError):
                        start_time = time(8, 0)
                        end_time = time(16, 0)

                    new_mission = Mission(
                        id=_uuid.uuid4(),
                        tenant_id=tenant.id,
                        name=f"{mt.name} - {current_date.isoformat()}",
                        mission_type_id=mt.id,
                        schedule_window_id=schedule_window.id if schedule_window else None,
                        date=current_date,
                        start_time=start_time,
                        end_time=end_time,
                        status="planned",
                        slots=mt.default_slots if hasattr(mt, 'default_slots') and mt.default_slots else [],
                    )
                    db.add(new_mission)
                    created_count += 1
                    missions_created_for_date = True

        # If no structured sections found, still count the date
        if not missions_created_for_date:
            skipped_count += 1

        current_date += timedelta(days=1)

    await db.commit()

    return {
        "template_id": str(tmpl.id),
        "template_name": tmpl.name,
        "date_from": str(data.date_from),
        "date_to": str(data.date_to),
        "days_processed": delta_days,
        "missions_created": created_count,
        "days_skipped": skipped_count,
    }


# ═══════════════════════════════════════════
# Swap Request Validation (Spec 3.18)
# ═══════════════════════════════════════════

@router.post("/swap-requests/{sr_id}/validate")
async def validate_swap_request(
    sr_id: UUID, tenant: CurrentTenant, user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Validate a swap request: check conflicts for both requester (freed) and target (assigned)."""
    result = await db.execute(
        select(SwapRequest).where(SwapRequest.id == sr_id, SwapRequest.tenant_id == tenant.id)
    )
    sr = result.scalar_one_or_none()
    if not sr:
        raise HTTPException(status_code=404, detail="בקשת החלפה לא נמצאה")

    requester_conflicts: list[dict] = []
    target_conflicts: list[dict] = []
    future_impact: list[str] = []

    # Get the requester's assignment and its mission
    req_assign_result = await db.execute(
        select(MissionAssignment).where(MissionAssignment.id == sr.requester_assignment_id)
    )
    req_assignment = req_assign_result.scalar_one_or_none()
    if not req_assignment:
        raise HTTPException(status_code=404, detail="שיבוץ המבקש לא נמצא")

    req_mission_result = await db.execute(
        select(Mission).where(Mission.id == req_assignment.mission_id)
    )
    req_mission = req_mission_result.scalar_one_or_none()
    if not req_mission:
        raise HTTPException(status_code=404, detail="משימת המבקש לא נמצאה")

    # If there's a target employee, check if assigning them to the requester's mission causes conflicts
    if sr.target_employee_id:
        # Check target employee time conflicts on the requester's mission date
        target_conflict_result = await db.execute(
            select(MissionAssignment, Mission)
            .join(Mission, MissionAssignment.mission_id == Mission.id)
            .where(
                MissionAssignment.employee_id == sr.target_employee_id,
                Mission.date == req_mission.date,
                Mission.id != req_mission.id,
                MissionAssignment.status != "replaced",
            )
        )
        for ma, existing_m in target_conflict_result.all():
            if existing_m.start_time < req_mission.end_time and existing_m.end_time > req_mission.start_time:
                target_conflicts.append({
                    "type": "time_overlap",
                    "mission_id": str(existing_m.id),
                    "mission_name": existing_m.name,
                    "time": f"{existing_m.start_time}-{existing_m.end_time}",
                })

        # Check rest hours for target
        yesterday = req_mission.date - timedelta(days=1)
        target_recent = await db.execute(
            select(Mission)
            .join(MissionAssignment, MissionAssignment.mission_id == Mission.id)
            .where(
                MissionAssignment.employee_id == sr.target_employee_id,
                Mission.date.in_([yesterday, req_mission.date]),
                Mission.id != req_mission.id,
                MissionAssignment.status != "replaced",
            )
            .order_by(Mission.date.desc(), Mission.end_time.desc())
        )
        recent = target_recent.scalars().all()
        if recent:
            last = recent[0]
            last_end = datetime.combine(last.date, last.end_time)
            this_start = datetime.combine(req_mission.date, req_mission.start_time)
            hours_rest = (this_start - last_end).total_seconds() / 3600
            if hours_rest < 16:
                target_conflicts.append({
                    "type": "insufficient_rest",
                    "hours_rest": round(hours_rest, 1),
                    "minimum_required": 16,
                })

    # For swap type, check if requester being freed creates coverage gaps
    if sr.swap_type == "swap" and sr.target_assignment_id:
        # Get target's assignment and mission
        tgt_assign_result = await db.execute(
            select(MissionAssignment).where(MissionAssignment.id == sr.target_assignment_id)
        )
        tgt_assignment = tgt_assign_result.scalar_one_or_none()
        if tgt_assignment:
            tgt_mission_result = await db.execute(
                select(Mission).where(Mission.id == tgt_assignment.mission_id)
            )
            tgt_mission = tgt_mission_result.scalar_one_or_none()
            if tgt_mission:
                # Check requester conflicts on target's mission
                req_conflict_result = await db.execute(
                    select(MissionAssignment, Mission)
                    .join(Mission, MissionAssignment.mission_id == Mission.id)
                    .where(
                        MissionAssignment.employee_id == sr.requester_employee_id,
                        Mission.date == tgt_mission.date,
                        Mission.id != tgt_mission.id,
                        MissionAssignment.status != "replaced",
                    )
                )
                for ma, existing_m in req_conflict_result.all():
                    if existing_m.start_time < tgt_mission.end_time and existing_m.end_time > tgt_mission.start_time:
                        requester_conflicts.append({
                            "type": "time_overlap",
                            "mission_id": str(existing_m.id),
                            "mission_name": existing_m.name,
                            "time": f"{existing_m.start_time}-{existing_m.end_time}",
                        })

    # Check future impact — count upcoming missions for both employees in next 7 days
    next_week = req_mission.date + timedelta(days=7)
    if sr.target_employee_id:
        target_upcoming = (await db.execute(
            select(func.count())
            .select_from(MissionAssignment)
            .join(Mission, MissionAssignment.mission_id == Mission.id)
            .where(
                MissionAssignment.employee_id == sr.target_employee_id,
                Mission.date >= req_mission.date,
                Mission.date <= next_week,
                MissionAssignment.status != "replaced",
            )
        )).scalar() or 0
        if target_upcoming >= 5:
            future_impact.append(f"לעובד המחליף {target_upcoming} משימות ב-7 ימים הקרובים")

    requester_upcoming = (await db.execute(
        select(func.count())
        .select_from(MissionAssignment)
        .join(Mission, MissionAssignment.mission_id == Mission.id)
        .where(
            MissionAssignment.employee_id == sr.requester_employee_id,
            Mission.date >= req_mission.date,
            Mission.date <= next_week,
            MissionAssignment.status != "replaced",
        )
    )).scalar() or 0
    if requester_upcoming <= 1:
        future_impact.append("למבקש כמעט אין משימות השבוע — החלפה עלולה להגדיל פערי עומס")

    is_fully_valid = len(requester_conflicts) == 0 and len(target_conflicts) == 0

    validation_result = {
        "requester_conflicts": requester_conflicts,
        "target_conflicts": target_conflicts,
        "future_impact": future_impact,
        "is_fully_valid": is_fully_valid,
    }

    # Persist validation result on the swap request
    sr.validation_result = validation_result
    await db.commit()

    return validation_result


# ═══════════════════════════════════════════
# Employee Import Preview & Import (Spec 3.13)
# ═══════════════════════════════════════════

@router.post("/schedule-windows/{window_id}/import-employees/preview", dependencies=[Depends(require_permission("missions", "write"))])
async def import_employees_preview(
    window_id: UUID,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
) -> dict:
    """Parse CSV/Excel file and return preview with per-row validation errors."""
    import csv
    import io

    # Validate window exists
    w_result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    if not w_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")

    content = await file.read()
    filename = file.filename or ""

    rows: list[dict] = []

    if filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row) if i < len(headers)})
        except ImportError:
            raise HTTPException(status_code=400, detail="ייבוא Excel לא נתמך. השתמש בקובץ CSV")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"שגיאה בקריאת קובץ Excel: {str(e)}")
    else:
        # Treat as CSV
        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append({k.strip(): v.strip() for k, v in row.items() if k})
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"שגיאה בקריאת קובץ CSV: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="הקובץ ריק או לא תקין")

    # Pre-load existing employees for matching
    emp_result = await db.execute(
        select(Employee).where(Employee.tenant_id == tenant.id)
    )
    existing_employees = emp_result.scalars().all()
    emp_by_number = {e.employee_number: e for e in existing_employees if e.employee_number}
    emp_by_name = {e.full_name: e for e in existing_employees}

    # Already in window
    existing_in_window = await db.execute(
        select(ScheduleWindowEmployee.employee_id).where(
            ScheduleWindowEmployee.schedule_window_id == window_id
        )
    )
    window_emp_ids = {row[0] for row in existing_in_window.all()}

    preview_rows = []
    valid_count = 0
    error_count = 0

    for idx, row in enumerate(rows, start=1):
        errors: list[str] = []
        employee_id = None
        employee_name = None
        status_str = "new"

        # Try to match by employee_number or full_name
        emp_number = row.get("employee_number") or row.get("מספר_עובד") or row.get("מספר עובד") or ""
        full_name = row.get("full_name") or row.get("שם_מלא") or row.get("שם מלא") or ""

        matched_emp = None
        if emp_number and emp_number in emp_by_number:
            matched_emp = emp_by_number[emp_number]
        elif full_name and full_name in emp_by_name:
            matched_emp = emp_by_name[full_name]

        if matched_emp:
            employee_id = str(matched_emp.id)
            employee_name = matched_emp.full_name
            if matched_emp.id in window_emp_ids:
                status_str = "already_exists"
                errors.append("העובד כבר משויך ללוח העבודה")
            elif not matched_emp.is_active:
                errors.append("העובד לא פעיל")
        else:
            if not emp_number and not full_name:
                errors.append("חסר מספר עובד או שם מלא")
            else:
                errors.append("עובד לא נמצא במערכת")

        if errors:
            error_count += 1
        else:
            valid_count += 1

        preview_rows.append({
            "row_number": idx,
            "employee_number": emp_number,
            "full_name": full_name or (employee_name or ""),
            "employee_id": employee_id,
            "status": status_str,
            "errors": errors,
        })

    return {
        "total_rows": len(rows),
        "valid_count": valid_count,
        "error_count": error_count,
        "rows": preview_rows,
    }


@router.post("/schedule-windows/{window_id}/import-employees", dependencies=[Depends(require_permission("missions", "write"))])
async def import_employees(
    window_id: UUID,
    tenant: CurrentTenant,
    user: CurrentUser,
    request: Request,
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
) -> dict:
    """Actually import employees from CSV/Excel into a schedule window (after preview)."""
    import csv
    import io

    # Validate window exists
    w_result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    window = w_result.scalar_one_or_none()
    if not window:
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")

    content = await file.read()
    filename = file.filename or ""
    rows: list[dict] = []

    if filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row) if i < len(headers)})
        except ImportError:
            raise HTTPException(status_code=400, detail="ייבוא Excel לא נתמך. השתמש בקובץ CSV")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"שגיאה בקריאת קובץ Excel: {str(e)}")
    else:
        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append({k.strip(): v.strip() for k, v in row.items() if k})
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"שגיאה בקריאת קובץ CSV: {str(e)}")

    # Pre-load employees
    emp_result = await db.execute(
        select(Employee).where(Employee.tenant_id == tenant.id, Employee.is_active.is_(True))
    )
    existing_employees = emp_result.scalars().all()
    emp_by_number = {e.employee_number: e for e in existing_employees if e.employee_number}
    emp_by_name = {e.full_name: e for e in existing_employees}

    existing_in_window = await db.execute(
        select(ScheduleWindowEmployee.employee_id).where(
            ScheduleWindowEmployee.schedule_window_id == window_id
        )
    )
    window_emp_ids = {row[0] for row in existing_in_window.all()}

    added = 0
    skipped = 0
    errors_list: list[dict] = []

    for idx, row in enumerate(rows, start=1):
        emp_number = row.get("employee_number") or row.get("מספר_עובד") or row.get("מספר עובד") or ""
        full_name = row.get("full_name") or row.get("שם_מלא") or row.get("שם מלא") or ""

        matched_emp = None
        if emp_number and emp_number in emp_by_number:
            matched_emp = emp_by_number[emp_number]
        elif full_name and full_name in emp_by_name:
            matched_emp = emp_by_name[full_name]

        if not matched_emp:
            errors_list.append({"row": idx, "error": "עובד לא נמצא"})
            skipped += 1
            continue

        if matched_emp.id in window_emp_ids:
            skipped += 1
            continue

        db.add(ScheduleWindowEmployee(
            schedule_window_id=window_id,
            employee_id=matched_emp.id,
        ))
        window_emp_ids.add(matched_emp.id)
        added += 1

    if added > 0:
        db.add(AuditLog(
            tenant_id=tenant.id, user_id=user.id, action="import_employees",
            entity_type="schedule_window", entity_id=window_id,
            after_state={"added": added, "skipped": skipped},
            ip_address=getattr(request.state, "real_ip", request.client.host if request.client else None),
        ))
        await db.commit()

    return {
        "added": added,
        "skipped": skipped,
        "errors": errors_list,
        "total_rows": len(rows),
    }


# ═══════════════════════════════════════════
# PDF Schedule Export (Spec Section 21)
# ═══════════════════════════════════════════

class PDFExportRequest(PydanticBaseModel):
    date: date
    template_id: UUID | None = None


@router.post("/schedule-windows/{window_id}/export-pdf", dependencies=[Depends(require_permission("missions", "read"))])
async def export_schedule_pdf(
    window_id: UUID,
    data: PDFExportRequest,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> FastAPIResponse:
    """Export schedule for a date as PDF with WeasyPrint."""
    from app.services.pdf_export import generate_schedule_pdf
    from app.models.tenant import Tenant

    # Validate window
    w_result = await db.execute(
        select(ScheduleWindow).where(ScheduleWindow.id == window_id, ScheduleWindow.tenant_id == tenant.id)
    )
    if not w_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="לוח עבודה לא נמצא")

    # Get tenant name
    tenant_result = await db.execute(select(Tenant).where(Tenant.id == tenant.id))
    tenant_obj = tenant_result.scalar_one_or_none()
    tenant_name = tenant_obj.name if tenant_obj else "—"

    # Get template if specified
    template_data = None
    if data.template_id:
        tmpl_result = await db.execute(
            select(DailyBoardTemplate).where(
                DailyBoardTemplate.id == data.template_id,
                DailyBoardTemplate.tenant_id == tenant.id,
            )
        )
        tmpl = tmpl_result.scalar_one_or_none()
        if tmpl:
            template_data = {"name": tmpl.name, "layout": tmpl.layout, "columns": tmpl.columns}

    # Fetch missions for date in this window
    missions_result = await db.execute(
        select(Mission).where(
            Mission.schedule_window_id == window_id,
            Mission.date == data.date,
            Mission.status.not_in(["cancelled"]),
        ).order_by(Mission.start_time)
    )
    missions = missions_result.scalars().all()

    # Build mission data with assignments and type names
    mission_dicts = []
    for m in missions:
        mt_result = await db.execute(select(MissionType).where(MissionType.id == m.mission_type_id))
        mt = mt_result.scalar_one_or_none()

        assign_result = await db.execute(
            select(MissionAssignment, Employee)
            .join(Employee, MissionAssignment.employee_id == Employee.id)
            .where(
                MissionAssignment.mission_id == m.id,
                MissionAssignment.status != "replaced",
            )
        )
        assignments = [
            {"employee_name": emp.full_name, "slot_id": ma.slot_id}
            for ma, emp in assign_result.all()
        ]

        mission_dicts.append({
            "name": m.name,
            "mission_type_name": mt.name if mt else "—",
            "start_time": str(m.start_time),
            "end_time": str(m.end_time),
            "status": m.status,
            "assignments": assignments,
        })

    pdf_bytes = generate_schedule_pdf(
        missions=mission_dicts,
        template=template_data,
        tenant_name=tenant_name,
        target_date=data.date,
    )

    return FastAPIResponse(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="schedule_{data.date.isoformat()}.pdf"',
        },
    )


# ═══════════════════════════════════════════
# Post-Mission Rules — Follow-up Configuration
# ═══════════════════════════════════════════

@router.get("/missions/{mission_id}/follow-ups")
async def get_follow_up_missions(
    mission_id: str, tenant: CurrentTenant, db: AsyncSession = Depends(get_db),
):
    """Get all follow-up missions linked to a parent mission."""
    result = await db.execute(
        select(Mission).where(
            Mission.parent_mission_id == mission_id,
            Mission.tenant_id == tenant.id,
        ).order_by(Mission.date, Mission.start_time)
    )
    follow_ups = result.scalars().all()
    return [
        {
            "id": str(m.id),
            "name": m.name,
            "date": str(m.date),
            "start_time": str(m.start_time),
            "end_time": str(m.end_time),
            "status": m.status,
            "is_activated": m.is_activated,
            "post_mission_config": m.post_mission_config,
            "assignments": [
                {
                    "id": str(a.id),
                    "employee_id": str(a.employee_id),
                    "slot_id": a.slot_id,
                    "status": a.status,
                }
                for a in (m.assignments or [])
            ],
        }
        for m in follow_ups
    ]


@router.get("/missions/{mission_id}/chain")
async def get_mission_chain(
    mission_id: str, tenant: CurrentTenant, db: AsyncSession = Depends(get_db),
):
    """Get the full mission chain — parent + all follow-ups recursively.
    Useful to see: סיור → כוננות → (next follow-up if any)."""
    # Find root
    current_result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    current = current_result.scalar_one_or_none()
    if not current:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")

    # Walk up to root
    root = current
    while root.parent_mission_id:
        parent_result = await db.execute(
            select(Mission).where(Mission.id == root.parent_mission_id)
        )
        parent = parent_result.scalar_one_or_none()
        if not parent:
            break
        root = parent

    # Walk down from root collecting chain
    chain = []

    async def collect_chain(mission):
        chain.append({
            "id": str(mission.id),
            "name": mission.name,
            "date": str(mission.date),
            "start_time": str(mission.start_time),
            "end_time": str(mission.end_time),
            "status": mission.status,
            "is_activated": mission.is_activated,
            "parent_mission_id": str(mission.parent_mission_id) if mission.parent_mission_id else None,
            "post_mission_config": mission.post_mission_config,
            "assignments_count": len(mission.assignments or []),
        })
        children_result = await db.execute(
            select(Mission).where(Mission.parent_mission_id == mission.id)
        )
        for child in children_result.scalars().all():
            await collect_chain(child)

    await collect_chain(root)
    return {"chain": chain, "root_id": str(root.id), "total": len(chain)}


@router.patch("/missions/{mission_id}/post-mission-config")
async def update_post_mission_config(
    mission_id: str,
    data: dict,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    """Update post-mission configuration for a specific mission.
    Allows overriding the auto-assign crew behavior, condition, etc."""
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    mission = result.scalar_one_or_none()
    if not mission:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")

    mission.post_mission_config = {
        **(mission.post_mission_config or {}),
        **data,
    }
    await db.commit()
    return {"status": "ok", "post_mission_config": mission.post_mission_config}


@router.post("/missions/{mission_id}/reassign-from-parent",
             dependencies=[Depends(require_permission("missions", "write"))])
async def reassign_from_parent(
    mission_id: str,
    tenant: CurrentTenant,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    """Re-copy assignments from parent mission to this follow-up mission.
    Useful when parent assignments change and you want to sync."""
    result = await db.execute(
        select(Mission).where(Mission.id == mission_id, Mission.tenant_id == tenant.id)
    )
    mission = result.scalar_one_or_none()
    if not mission:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    if not mission.parent_mission_id:
        raise HTTPException(status_code=400, detail="משימה זו אינה משימת המשך — אין משימת אב")

    # Get parent assignments
    parent_assignments_result = await db.execute(
        select(MissionAssignment).where(
            MissionAssignment.mission_id == mission.parent_mission_id,
            MissionAssignment.status != "replaced",
        )
    )
    parent_assignments = parent_assignments_result.scalars().all()

    # Get follow-up mission type for slot matching
    mt_result = await db.execute(select(MissionType).where(MissionType.id == mission.mission_type_id))
    followup_mt = mt_result.scalar_one_or_none()
    followup_slots = (followup_mt.required_slots or []) if followup_mt else []

    # Clear existing assignments on follow-up
    existing_result = await db.execute(
        select(MissionAssignment).where(MissionAssignment.mission_id == mission.id)
    )
    for existing in existing_result.scalars().all():
        existing.status = "replaced"

    # Copy from parent
    copied = 0
    for pa in parent_assignments:
        target_slot = None
        for fs in followup_slots:
            if fs.get("work_role_id") == str(pa.work_role_id):
                target_slot = fs.get("slot_id", "default")
                break
        if not target_slot and followup_slots:
            target_slot = followup_slots[0].get("slot_id", "default")

        if target_slot:
            new_assignment = MissionAssignment(
                mission_id=mission.id,
                employee_id=pa.employee_id,
                work_role_id=pa.work_role_id,
                slot_id=target_slot,
                status="assigned",
            )
            db.add(new_assignment)
            copied += 1

    await db.commit()
    return {
        "status": "ok",
        "copied": copied,
        "from_parent": str(mission.parent_mission_id),
        "message": f"הועתקו {copied} שיבוצים ממשימת האב",
    }
